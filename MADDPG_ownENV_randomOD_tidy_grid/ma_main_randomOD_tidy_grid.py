import sys
# sys.path.append('F:\githubClone\Multi_agent_AAC\old_framework_test')
# sys.path.append('D:\Multi_agent_AAC\old_framework_test')
from openpyxl import load_workbook
from openpyxl import Workbook
import argparse
import datetime
import pandas as pd
import numpy as np
import torch
import os
import time
import matplotlib.animation as animation
import pickle
import wandb
from parameters_randomOD_tidy_grid import initialize_parameters
from maddpg_agent_randomOD_tidy_grid import MADDPG
from utils_randomOD_tidy_grid import *
from copy import deepcopy
import torch
import matplotlib.pyplot as plt
import matplotlib
from shapely.geometry import LineString, Point, Polygon
from shapely.strtree import STRtree
from matplotlib.markers import MarkerStyle
import math
from matplotlib.transforms import Affine2D
from Utilities_own_randomOD_tidy_grid import *
import csv

num_devices = torch.cuda.device_count()
print("Number of GPUs:", num_devices)
# Get the names of the available GPUs
gpu_names = [torch.cuda.get_device_name(i) for i in range(num_devices)]
print("GPU Names:", gpu_names)
if torch.cuda.is_available():
    device = torch.device('cuda')
    print('Using GPU')
else:
    device = torch.device('cpu')
    print('Using CPU')


def initialize_excel_file(file_path):
    # Create a new workbook and add three empty sheets
    wb = Workbook()
    wb.save(file_path)


def append_to_excel(file_path, data):
    # Try to load the workbook, and create it if it does not exist
    try:
        wb = load_workbook(file_path)
        new_workbook = False
    except FileNotFoundError:
        wb = Workbook()
        new_workbook = True
        # wb.remove(wb.active)  # Remove the default sheet created by openpyxl if not need
    except Exception as e:  # Catch other exceptions, such as invalid file format
        print(f"An error occurred: {e}")
        return
    if new_workbook:
        if 'Sheet' in wb.sheetnames:
            # If the sheet exists, remove it
            wb.remove('Sheet')
    # Check if the required sheets exist, if not create them
    if len(data) == 3:  # for record eps time
        number_of_sheets = len(data)
        required_sheets = ['Sheet'+str(i) for i in range(number_of_sheets)]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(title=sheet_name)

        # Append the data to the respective sheets
        for i, portion in enumerate(data):
            sheet = wb[required_sheets[i]]
            if isinstance(portion, list):
                for item in portion:
                    sheet.append(item if isinstance(item, list) else [item])
                # If this is the last portion of data, append -1 to the last row
                if i == len(data) - 1:
                    sheet.append([-1])
            else:
                sheet.append([portion])
    else:
        if isinstance(data[0], list):  # for record reward
            # In reward record, we are recording in the form that one step have 3 agents
            required_sheets = ["agent0", "agent1", "agent2"]
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(title=sheet_name)
            for idx, step_reward in enumerate(data):
                for agent_idx, single_sheet_name in enumerate(required_sheets):
                    sheet = wb[single_sheet_name]
                    sheet.append(step_reward[agent_idx])
                    if idx == len(data)-1:
                        sheet.append([-9999])

        else:  # for record noise
            wb.create_sheet(title='noise given at each step')
            sheet = wb['noise given at each step']
            for i, portion in enumerate(data):
                sheet.append(list(portion))
            sheet.append([-999])

    # Save the workbook
    wb.save(file_path)


def animate(frame_num, ax, env, trajectory_eachPlay):
    ax.clear()
    plt.axis('equal')
    plt.xlim(env.bound[0], env.bound[1])
    plt.ylim(env.bound[2], env.bound[3])
    plt.axvline(x=env.bound[0], c="green")
    plt.axvline(x=env.bound[1], c="green")
    plt.axhline(y=env.bound[2], c="green")
    plt.axhline(y=env.bound[3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")

    # draw occupied_poly
    for one_poly in env.world_map_2D_polyList[0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in env.world_map_2D_polyList[0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    for agentIdx, agent in env.all_agents.items():
        plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                 marker=MarkerStyle(">",
                                    fillstyle="right",
                                    transform=Affine2D().rotate_deg(math.degrees(agent.heading))),
                 color='y')
        plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
        plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='y', markersize=10)
        plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)

        # link individual drone's starting position with its goal
        ini = agent.ini_pos
        for wp in agent.goal:
            # plt.plot(wp[0], wp[1], marker='*', color='y', markersize=10)
            plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c')
            ini = wp

    for a_idx, agent in enumerate(trajectory_eachPlay[frame_num]):
        x, y = agent[0], agent[1]
        plt.plot(x, y, 'o', color='r')

        plt.text(x-1, y-1, 'agent_'+str(a_idx)+'_'+str(round(float(frame_num), 2)))

        self_circle = Point(x, y).buffer(env.all_agents[0].protectiveBound, cap_style='round')
        grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')
        ax.add_patch(grid_mat_Scir)

    return ax.patches + [ax.texts]


def main(args):

    if args.mode == "train":
        today = datetime.date.today()
        current_date = today.strftime("%d%m%y")
        current_time = datetime.datetime.now()
        formatted_time = current_time.strftime("%H_%M_%S")
        file_name = 'D:\MADDPG_2nd_jp/' + str(current_date) + '_' + str(formatted_time)
        if not os.path.exists(file_name):
            os.makedirs(file_name)
        plot_file_name = file_name + '/toplot'
        if not os.path.exists(plot_file_name):
            os.makedirs(plot_file_name)
        # ------------ this portion is to save using excel instead of pickle -----------
        # excel_file_path_reward = plot_file_name + '/all_episode_reward.xlsx'
        # excel_file_path_noise = plot_file_name + '/all_episode_noise.xlsx'
        # excel_file_path_time = plot_file_name + '/all_episode_time.xlsx'
        # initialize_excel_file(excel_file_path_reward)
        # initialize_excel_file(excel_file_path_noise)
        # initialize_excel_file(excel_file_path_time)
        # ------------ end of this portion is to save using excel instead of pickle -----------

    # use_wanDB = False
    use_wanDB = True
    # simply_view_evaluation = True  # True = don't save gif
    simply_view_evaluation = False  # True = don't save gif

    if use_wanDB:
        wandb.login(key="efb76db851374f93228250eda60639c70a93d1ec")
        wandb.init(
            # set the wandb project where this run will be logged
            project="MADDPG_FrameWork",
            name='MADDPG_test_'+str(current_date) + '_' + str(formatted_time),
            # track hyperparameters and run metadata
            config={
                # "learning_rate": args.a_lr,
                "epochs": args.max_episodes,
            }
        )

    # -------------- create my own environment -----------------
    n_episodes, max_t, eps_start, eps_end, eps_period, eps, env, \
    agent_grid_obs, BUFFER_SIZE, BATCH_SIZE, GAMMA, TAU, learning_rate, UPDATE_EVERY, seed_used, max_xy = initialize_parameters()
    # total_agentNum = len(pd.read_excel(env.agentConfig))
    total_agentNum = 3
    # UPDATE_EVERY = 30
    UPDATE_EVERY = 1
    # max_nei_num = 5
    # create world
    # actor_dim = [6+(total_agentNum-1)*2, 10, 6]  # dim host, maximum dim grid, dim other drones
    # critic_dim = [6+(total_agentNum-1)*2, 10, 6]
    actor_dim = [6, 9, 6]  # dim host, maximum dim grid, dim other drones
    # actor_dim = [9, 9, 9]  # dim host, maximum dim grid, dim other drones
    # actor_dim = [16, 9, 6]  # dim host, maximum dim grid, dim other drones
    critic_dim = [6, 9, 6]
    actor_hidden_state = 64
    # critic_dim = [9, 9, 9]
    # critic_dim = [16, 9, 6]
    n_actions = 2
    acc_range = [-4, 4]

    actorNet_lr = 0.001
    criticNet_lr = 0.001

    # noise parameter ini
    largest_Nsigma = 0.5
    smallest_Nsigma = 0.15
    ini_Nsigma = largest_Nsigma

    max_spd = 15
    env.create_world(total_agentNum, n_actions, GAMMA, TAU, UPDATE_EVERY, largest_Nsigma, smallest_Nsigma, ini_Nsigma, max_xy, max_spd, acc_range)

    # --------- my own -----------
    n_agents = len(env.all_agents)
    n_actions = n_actions

    torch.manual_seed(args.seed)  # this is the seed

    if args.algo == "maddpg":
        model = MADDPG(actor_dim, critic_dim, n_actions, actor_hidden_state, n_agents, args, criticNet_lr, actorNet_lr, GAMMA, TAU)

    episode = 0
    total_step = 0
    score_history = []
    eps_reward_record = []
    eps_check_collision = []
    eps_noise_record = []
    eps_end = 5000  # at eps = eps_end, the eps value drops to lowest value which is 0.03 (this value is fixed)
    noise_start_level = 1
    training_start_time = time.time()

    # ------------ record episode time ------------- #
    eps_time_record = []
    # ----------- record each collision checking version running time and decision -------#
    collision_count = 0

    if args.mode == "eval":
        # args.max_episodes = 1  # only evaluate one episode during evaluation mode.
        args.max_episodes = 100

    # while episode < args.max_episodes:
    while episode < args.max_episodes:  # start of an episode

        # ------------ my own env.reset() ------------ #
        episode_start_time = time.time()
        episode += 1
        eps_reset_start_time = time.time()
        cur_state, norm_cur_state = env.reset_world(total_agentNum, show=0)
        eps_reset_time_used = (time.time()-eps_reset_start_time)*1000
        print("current episode {} reset time used is {} milliseconds".format(episode, eps_reset_time_used))  # need to + 1 here, or else will misrecord as the previous episode
        step_collision_record = [[], [], []]  # reset at each episode, so that we can record down collision at each step for each agent.
        eps_status_holder = [None] * n_agents
        episode_decision = [False] * 2
        agents_added = []
        eps_reward = []
        eps_noise = []
        step_time_breakdown = []
        # print("current episode is {}, scaling factor is {}".format(episode, model.var[0]))
        step = 0
        agent_added = 0  # this is an initialization for each new episode
        accum_reward = 0

        trajectory_eachPlay = []

        pre_fix = r'D:\MADDPG_2nd_jp\141223_13_51_10\interval_record_eps'
        episode_to_check = str(14000)
        load_filepath_0 = pre_fix + '\episode_' + episode_to_check + '_agent_0actor_net.pth'
        load_filepath_1 = pre_fix + '\episode_' + episode_to_check + '_agent_1actor_net.pth'
        load_filepath_2 = pre_fix + '\episode_' + episode_to_check + '_agent_2actor_net.pth'
        # load_filepath_3 = pre_fix + '\episode_' + episode_to_check + '_agent_3actor_net.pth'
        # load_filepath_4 = pre_fix + '\episode_' + episode_to_check + '_agent_4actor_net.pth'

        if args.mode == "eval":
            # model.load_model([load_filepath_0, load_filepath_1, load_filepath_2, load_filepath_3, load_filepath_4])
            model.load_model([load_filepath_0, load_filepath_1, load_filepath_2])
        while True:  # start of an episode (this episode ends when (agent_added < max_agent_to_add))
            if args.mode == "train":
                step_start_time = time.time()
                step_reward_record = [None] * n_agents

                # cur_state, norm_cur_state = env.fill_agent_reset(cur_state, norm_cur_state, agents_added)  # if a new agent is filled, we need to reset the state information for the newly added agents

                step_obtain_action_time_start = time.time()
                action, step_noise_val = model.choose_action(norm_cur_state, total_step, episode, step, eps_end, noise_start_level, noisy=False) # noisy is false because we are using stochastic policy
                # action, step_noise_val, agents_hn = model.choose_action(norm_cur_state, total_step, episode, step, eps_end, noise_start_level, noisy=False) # noisy is false because we are using stochastic policy
                generate_action_time = (time.time() - step_obtain_action_time_start)*1000
                # print("current step obtain action time used is {} milliseconds".format(generate_action_time))

                # action = model.choose_action(cur_state, episode, noisy=True)

                one_step_transition_start = time.time()
                next_state, norm_next_state = env.step(action, step)
                step_transition_time = (time.time() - one_step_transition_start)*1000
                # print("current step transition time used is {} milliseconds".format(step_transition_time))

                # reward_aft_action, done_aft_action, check_goal, step_reward_record, agents_added = env.get_step_reward_5_v3(step, step_reward_record)   # remove reached agent here
                # reward_aft_action, done_aft_action, check_goal, step_reward_record = env.get_step_reward_5_v3(step, step_reward_record)   # remove reached agent here

                one_step_reward_start = time.time()
                reward_aft_action, done_aft_action, check_goal, step_reward_record, status_holder, step_collision_record = env.ss_reward(step, step_reward_record, eps_status_holder, step_collision_record)   # remove reached agent here
                reward_generation_time = (time.time() - one_step_reward_start)*1000
                # print("current step reward time used is {} milliseconds".format(reward_generation_time))

                # reward_aft_action = [eachRWD / 300 for eachRWD in reward_aft_action]  # scale the reward down
                # new_length = len(agents_added)  # check if length of agnet_to_remove increased during each step
                # agent_added = agent_added + new_length

                step += 1  # current play step
                total_step += 1  # steps taken from 1st episode
                eps_noise.append(step_noise_val)
                if args.algo == "maddpg" or args.algo == "commnet":
                    obs = []
                    next_obs = []
                    # ------------- to store norm or non-norm state into experience replay ---------------
                    for elementIdx, element in enumerate(norm_cur_state):
                    # for elementIdx, element in enumerate(cur_state):
                        if elementIdx != len(norm_cur_state)-1:  # meaning is not the last element
                        # if elementIdx != len(cur_state)-1:  # meaning is not the last element
                            obs.append(torch.from_numpy(np.stack(element)).data.float().to(device))
                        else:
                            sur_agents = []
                            for each_agent_list in element:
                                sur_agents.append(torch.from_numpy(np.squeeze(np.array(each_agent_list), axis=1)).float())
                            obs.append(sur_agents)

                    for elementIdx, element in enumerate(norm_next_state):
                    # for elementIdx, element in enumerate(cur_state):
                        if elementIdx != len(norm_next_state)-1:  # meaning is not the last element
                        # if elementIdx != len(cur_state)-1:  # meaning is not the last element
                            next_obs.append(torch.from_numpy(np.stack(element)).data.float().to(device))
                        else:
                            sur_agents = []
                            for each_agent_list in element:
                                sur_agents.append(torch.from_numpy(np.squeeze(np.array(each_agent_list), axis=1)).float())
                            next_obs.append(sur_agents)
                    # ------------------ end of store norm or non-norm state into experience replay --------------------
                    rw_tensor = torch.FloatTensor(np.array(reward_aft_action)).to(device)
                    ac_tensor = torch.FloatTensor(action).to(device)
                    done_tensor = torch.FloatTensor(done_aft_action).to(device)
                    if args.algo == "commnet" and next_obs is not None:
                        model.memory.push(obs.data, ac_tensor, next_obs, rw_tensor, done_tensor)
                    if args.algo == "maddpg":
                        model.memory.push(obs, ac_tensor, next_obs, rw_tensor, done_tensor)
                else:
                    model.memory(cur_state, action, reward_aft_action, next_state, done_aft_action)

                # accum_reward = accum_reward + reward_aft_action[0]  # we just take the first agent's reward, because we are using a joint reward, so all agents obtain the same reward.
                accum_reward = accum_reward + sum(reward_aft_action)

                step_update_time_start = time.time()
                c_loss, a_loss = model.update_myown(episode, total_step, UPDATE_EVERY, wandb)  # last working learning framework
                update_time_used = (time.time() - step_update_time_start)*1000
                print("current step update time used is {} milliseconds".format(update_time_used))
                cur_state = next_state
                norm_cur_state = norm_next_state
                eps_reward.append(step_reward_record)
                whole_step_time = (time.time()-step_start_time)*1000
                print("current episode, one whole step time used is {} milliseconds".format(whole_step_time))
                step_time_breakdown.append([generate_action_time, step_transition_time, reward_generation_time,
                                            update_time_used, whole_step_time])
                if ((args.episode_length < step) and (300 not in reward_aft_action)):
                    episode_decision[0] = True
                    print("Agents stuck in some places, maximum step in one episode reached, current episode {} ends, all {} steps used".format(episode, args.episode_length))
                elif (True in done_aft_action):
                    episode_decision[1] = True
                    print("Some agent triggers termination condition like collision, current episode {} ends at step {}".format(episode, step-1))  # we need to -1 here, because we perform step + 1 after each complete step. Just to be consistent with the step count inside the reward function.
                # elif (agent_added>50):
                #     episode_decision[2] = True
                #     print("More than 50 drones has reaches the destination, current episode {} ends".format(episode))
                # if ((args.episode_length < step) and (300 not in reward_aft_action)) or (True in done_aft_action) or (agent_added>50):

                if True in episode_decision:
                    # end of an episode starts here

                    # time_used = time.time() - start_time
                    # print("update function used {} seconds to run".format(time_used))
                    # here onwards is end of an episode's play
                    score_history.append(accum_reward)

                    # print("[Episode %05d] reward %6.4f time used is %.2f sec" % (episode, accum_reward, time_used))
                    print("[Episode %05d] reward %6.4f" % (episode, accum_reward))

                    if use_wanDB:
                        wandb.log({'overall_reward': float(accum_reward)})
                    # if c_loss and a_loss:
                    #     for idx, val in enumerate(c_loss):
                    #         print(" agent %s, a_loss %3.2f c_loss %3.2f" % (idx, a_loss[idx].item(), c_loss[idx].item()))
                            # wandb.log({'agent' + str(idx) + 'actor_loss': float(a_loss[idx].item())})
                            # wandb.log({'agent' + str(idx) + 'critic_loss': float(c_loss[idx].item())})
                    if episode % args.save_interval == 0 and args.mode == "train":
                        save_model = time.time()
                        # save the models at a predefined interval
                        # save model to my own directory
                        filepath = file_name+'/interval_record_eps'
                        model.save_model(episode, filepath)  # this is the original save model
                        time_used_for_csv_model_save = (time.time()-save_model)*1000  # *1000 for milliseconds
                        print("current episode used time in save csv and model is {} milliseconds".format(episode, time_used_for_csv_model_save))
                    # save episodes reward for entire system at each of one episode
                    eps_reward_record.append(eps_reward)
                    eps_check_collision.append(step_collision_record)
                    eps_noise_record.append(eps_noise)
                    epsTime = time.time()-episode_start_time
                    eps_time_record.append([eps_reset_time_used, epsTime, step_time_breakdown])
                    print("episode {} used time in calculation is  {} seconds".format(episode, epsTime))


                    # --------- removed to save time ----------
                    # storage_time = time.time()  # storage time is too long, one episode is >= 150 milliseconds
                    # append_to_excel(excel_file_path_time, eps_time_record[-1])
                    # append_to_excel(excel_file_path_noise, eps_noise_record[-1])
                    # append_to_excel(excel_file_path_reward, eps_reward_record[-1])
                    # end_of_storage_time = (time.time()-storage_time)*1000  # just compute storage time in milliseconds
                    # print("episode {} used time in storage is  {} milliseconds".format(episode, end_of_storage_time))
                    # total_time_one_episode = (end_of_storage_time)/1000 + epsTime
                    # print("episode {} used time in total {} seconds".format(episode, total_time_one_episode))
                    #

                    break  # this is to break out from "while True:", which is one play
            elif args.mode == "eval":
                step_reward_record = [None] * n_agents
                action, step_noise_val = model.choose_action(norm_cur_state, total_step, episode, step, eps_end, noise_start_level, noisy=False)
                # action = model.choose_action(cur_state, episode, noisy=False)
                # action = env.get_actions_noCR()  # only update heading, don't update any other attribute
                # for a_idx, action_ele in enumerate(action):
                #     action[a_idx] = [-0.3535, 0.3535]
                next_state, norm_next_state = env.step(action, step)  # no heading update here
                reward_aft_action, done_aft_action, check_goal, step_reward_record, eps_status_holder, step_collision_record = env.ss_reward(step, step_reward_record, eps_status_holder, step_collision_record)
                # reward_aft_action, done_aft_action, check_goal, step_reward_record = env.get_step_reward_5_v3(step, step_reward_record)

                step += 1
                total_step += 1
                cur_state = next_state
                norm_cur_state = norm_next_state
                trajectory_eachPlay.append([[each_agent_traj[0], each_agent_traj[1]] for each_agent_traj in cur_state[0]])
                accum_reward = accum_reward + sum(reward_aft_action)


                if args.episode_length < step or (True in done_aft_action):  # when termination condition reached
                # if args.episode_length < step:  # when termination condition reached, without counting drone collision to buildings/wall
                    # display current episode out status through status_holder
                    for each_agent_idx, each_agent in enumerate(eps_status_holder):
                        for step_idx, step_reward_decomposition in enumerate(each_agent):
                            print(r"agent {}, step {}, distance to goal is {} m, goal reward is {}, ref line reward is {}, current step reward is {}.".format(each_agent_idx, step_idx, step_reward_decomposition[0], step_reward_decomposition[1], step_reward_decomposition[2], step_reward_decomposition[3]))
                            print("near goal reward is {}".format(step_reward_decomposition[6]))
                            print("current spd is {} m/s, curent spd penalty is {}". format(step_reward_decomposition[5], step_reward_decomposition[4]))
                    print("[Episode %05d] reward %6.4f " % (episode, accum_reward))

                    # if simply_view_evaluation:
                    # # ------------------ static display trajectory ---------------------------- #
                    #     os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
                    #     matplotlib.use('TkAgg')
                    #     fig, ax = plt.subplots(1, 1)
                    #     # display initial condition
                    #     # global_state = env.reset_world(show=0)  # just a dummy to reset all condition so that initial condition can be added to the output graph
                    #     for agentIdx, agent in env.all_agents.items():
                    #         plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                    #                  marker=MarkerStyle(">",
                    #                                     fillstyle="right",
                    #                                     transform=Affine2D().rotate_deg(math.degrees(agent.heading))),
                    #                  color='y')
                    #         plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
                    #         # plot self_circle of the drone
                    #         self_circle = Point(agent.ini_pos[0],
                    #                             agent.ini_pos[1]).buffer(agent.protectiveBound, cap_style='round')
                    #         grid_mat_Scir = shapelypoly_to_matpoly(self_circle, inFill=False, Edgecolor='k')
                    #         ax.add_patch(grid_mat_Scir)
                    #
                    #         # plot drone's detection range
                    #         detec_circle = Point(agent.ini_pos[0],
                    #                              agent.ini_pos[1]).buffer(agent.detectionRange / 2, cap_style='round')
                    #         detec_circle_mat = shapelypoly_to_matpoly(detec_circle, inFill=False, Edgecolor='g')
                    #         ax.add_patch(detec_circle_mat)
                    #
                    #         # link individual drone's starting position with its goal
                    #         ini = agent.ini_pos
                    #         for wp in agent.goal:
                    #             plt.plot(wp[0], wp[1], marker='*', color='y', markersize=10)
                    #             plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c')
                    #             ini = wp
                    #         plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='y', markersize=10)
                    #         plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)
                    #
                    #     # draw trajectory in current episode
                    #     for trajectory_idx, trajectory_val in enumerate(trajectory_eachPlay):  # each time step
                    #         for agentIDX, each_agent_traj in enumerate(trajectory_val):  # for each agent's motion in a time step
                    #             x, y = each_agent_traj[0], each_agent_traj[1]
                    #             plt.plot(x, y, 'o', color='r')
                    #
                    #             # plt.text(x-1, y-1, str(round(float(reward_each_agent[trajectory_idx][agentIDX]),2)))
                    #
                    #             self_circle = Point(x, y).buffer(env.all_agents[0].protectiveBound, cap_style='round')
                    #             grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')
                    #             ax.add_patch(grid_mat_Scir)
                    #
                    #     # draw occupied_poly
                    #     for one_poly in env.world_map_2D_polyList[0][0]:
                    #         one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
                    #         ax.add_patch(one_poly_mat)
                    #     # draw non-occupied_poly
                    #     for zero_poly in env.world_map_2D_polyList[0][1]:
                    #         zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
                    #         # ax.add_patch(zero_poly_mat)
                    #
                    #     # show building obstacles
                    #     for poly in env.buildingPolygons:
                    #         matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
                    #         ax.add_patch(matp_poly)
                    #
                    #     plt.axis('equal')
                    #     plt.xlim(env.bound[0], env.bound[1])
                    #     plt.ylim(env.bound[2], env.bound[3])
                    #     plt.axvline(x=env.bound[0], c="green")
                    #     plt.axvline(x=env.bound[1], c="green")
                    #     plt.axhline(y=env.bound[2], c="green")
                    #     plt.axhline(y=env.bound[3], c="green")
                    #     plt.xlabel("X axis")
                    #     plt.ylabel("Y axis")
                    #     plt.show()
                    # # ------------------ end of static display trajectory ---------------------------- #
                    #
                    # # ---------- new save as gif ----------------------- #
                    # else:
                    #     os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
                    #     matplotlib.use('TkAgg')
                    #     fig, ax = plt.subplots(1, 1)
                    #
                    #     plt.axis('equal')
                    #     plt.xlim(env.bound[0], env.bound[1])
                    #     plt.ylim(env.bound[2], env.bound[3])
                    #     plt.axvline(x=env.bound[0], c="green")
                    #     plt.axvline(x=env.bound[1], c="green")
                    #     plt.axhline(y=env.bound[2], c="green")
                    #     plt.axhline(y=env.bound[3], c="green")
                    #     plt.xlabel("X axis")
                    #     plt.ylabel("Y axis")
                    #
                    #     # draw occupied_poly
                    #     for one_poly in env.world_map_2D_polyList[0][0]:
                    #         one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
                    #         ax.add_patch(one_poly_mat)
                    #     # draw non-occupied_poly
                    #     for zero_poly in env.world_map_2D_polyList[0][1]:
                    #         zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
                    #         # ax.add_patch(zero_poly_mat)
                    #
                    #     # show building obstacles
                    #     for poly in env.buildingPolygons:
                    #         matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
                    #         ax.add_patch(matp_poly)
                    #
                    #     for agentIdx, agent in env.all_agents.items():
                    #         plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                    #                  marker=MarkerStyle(">",
                    #                                     fillstyle="right",
                    #                                     transform=Affine2D().rotate_deg(math.degrees(agent.heading))),
                    #                  color='y')
                    #         plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
                    #         # plot self_circle of the drone
                    #         self_circle = Point(agent.ini_pos[0],
                    #                             agent.ini_pos[1]).buffer(agent.protectiveBound, cap_style='round')
                    #         grid_mat_Scir = shapelypoly_to_matpoly(self_circle, inFill=False, Edgecolor='k')
                    #         ax.add_patch(grid_mat_Scir)
                    #
                    #         # plot drone's detection range
                    #         detec_circle = Point(agent.ini_pos[0],
                    #                              agent.ini_pos[1]).buffer(agent.detectionRange / 2, cap_style='round')
                    #         detec_circle_mat = shapelypoly_to_matpoly(detec_circle, inFill=False, Edgecolor='g')
                    #         ax.add_patch(detec_circle_mat)
                    #
                    #         plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='y', markersize=10)
                    #         plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)
                    #
                    #     # Create animation
                    #     ani = animation.FuncAnimation(fig, animate, fargs=(ax, env, trajectory_eachPlay), frames=len(trajectory_eachPlay), interval=300, blit=False)
                    #     # Save as GIF
                    #     gif_path = pre_fix + '\episode_' + episode_to_check + 'simulation_num_'+str(episode)+'.gif'
                    #     ani.save(gif_path, writer='pillow')
                    #
                    #     # Close figure
                    #     plt.close(fig)
                    if True in done_aft_action and step < args.episode_length:
                        collision_count = collision_count + 1
                    break

    if args.mode == "train":  # only save pickle at end of training to save computational time.
        with open(plot_file_name + '/all_episode_reward.pickle', 'wb') as handle:
            pickle.dump(eps_reward_record, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(plot_file_name + '/all_episode_noise.pickle', 'wb') as handle:
            pickle.dump(eps_noise_record, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(plot_file_name + '/all_episode_time.pickle', 'wb') as handle:
            pickle.dump(eps_time_record, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(plot_file_name + '/all_episode_collision.pickle', 'wb') as handle:
            pickle.dump(eps_check_collision, handle, protocol=pickle.HIGHEST_PROTOCOL)
        with open(file_name + '/GFG.csv', 'w') as f:
            # using csv.writer method from CSV package
            write = csv.writer(f)
            write.writerows([score_history])
    else:
        print("total collision count is {}".format(collision_count))
    print(f'training finishes, time spent: {datetime.timedelta(seconds=int(time.time() - training_start_time))}')
    if use_wanDB:
        wandb.finish()


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--scenario', default="simple_spread", type=str)
    parser.add_argument('--max_episodes', default=35000, type=int)  # run for a total of 50000 episodes
    parser.add_argument('--algo', default="maddpg", type=str, help="commnet/bicnet/maddpg")
    parser.add_argument('--mode', default="train", type=str, help="train/eval")
    parser.add_argument('--episode_length', default=150, type=int)  # maximum play per episode
    parser.add_argument('--memory_length', default=int(1e5), type=int)
    parser.add_argument('--seed', default=777, type=int)  # may choose to use 3407
    parser.add_argument('--batch_size', default=512, type=int)  # original 512
    parser.add_argument('--render_flag', default=False, type=bool)
    parser.add_argument('--ou_theta', default=0.15, type=float)
    parser.add_argument('--ou_mu', default=0.0, type=float)
    parser.add_argument('--ou_sigma', default=0.2, type=float)
    parser.add_argument('--epsilon_decay', default=10000, type=int)
    parser.add_argument('--tensorboard', default=True, action="store_true")
    parser.add_argument("--save_interval", default=1000, type=int)  # save model for every 5000 episodes
    parser.add_argument("--model_episode", default=60000, type=int)
    parser.add_argument('--episode_before_train', default=10, type=int)  # original 1000
    parser.add_argument('--log_dir', default=datetime.datetime.now().strftime('%Y%m%d_%H%M%S'))

    args = parser.parse_args()

    main(args)
