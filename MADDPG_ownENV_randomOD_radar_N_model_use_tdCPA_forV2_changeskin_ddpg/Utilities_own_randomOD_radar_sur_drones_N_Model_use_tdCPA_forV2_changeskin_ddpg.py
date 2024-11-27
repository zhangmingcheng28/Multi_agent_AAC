# -*- coding: utf-8 -*-
"""
@Time    : 3/13/2023 1:28 PM
@Author  : Mingcheng
@FileName: 
@Description: 
@Package dependency:
"""
from matplotlib.patches import Polygon as matPolygon
import matplotlib.path as mpath
import matplotlib.patches as mpatches
from matplotlib.path import Path
from matplotlib.patches import PathPatch
from scipy.ndimage import gaussian_filter
from matplotlib.colors import LinearSegmentedColormap
import torch as T
import numpy as np
import torch
import os
import matplotlib
import matplotlib.pyplot as plt
from shapely.strtree import STRtree
from shapely.geometry import LineString, Point, Polygon
import matplotlib.colors as colors
from shapely.ops import nearest_points
from matplotlib.transforms import Affine2D
import matplotlib.animation as animation
from openpyxl import load_workbook
from openpyxl import Workbook
import cairosvg
import matplotlib.patches as patches
import io
from PIL import Image
from matplotlib.markers import MarkerStyle
import math


def obtain_euclidean_dist_list_all_AC(flight_ratio_per_eps_all_AC, trajectory_eachPlay, env):
    for agentIDX, agent in env.all_agents.items():
        if agent.reach_target == True:
            desired_distance = np.linalg.norm(np.array(agent.goal[-1]) - agent.ini_pos)
            each_ac_total_dist = 0
            for trajectory_idx in range(1, len(trajectory_eachPlay)):
                pre_x, pre_y = trajectory_eachPlay[trajectory_idx-1][agentIDX][0], trajectory_eachPlay[trajectory_idx-1][agentIDX][1]
                x, y = trajectory_eachPlay[trajectory_idx][agentIDX][0], trajectory_eachPlay[trajectory_idx][agentIDX][1]
                distance = np.linalg.norm(np.array([x, y]) - np.array([pre_x, pre_y]))
                each_ac_total_dist = each_ac_total_dist + distance
            first_pt = np.array([trajectory_eachPlay[0][agentIDX][0],
                                 trajectory_eachPlay[0][agentIDX][1]])
            dist_ini_pos_to_1stPt = np.linalg.norm(first_pt - agent.ini_pos)
            each_ac_total_dist = each_ac_total_dist + dist_ini_pos_to_1stPt + 5  # this 5 is because, when ac reached, just the protective zone reach is consider reached, so we add 5.
            flight_ratio_per_eps_all_AC.append(each_ac_total_dist/desired_distance)
    return flight_ratio_per_eps_all_AC


def plot_linestring(ax, linestring, color='black', zorder=0):
    x, y = linestring.xy
    ax.plot(x, y, color=color, linewidth=2)


def plot_bounding_box(ax, bbox, edgecolor='r', facecolor='none'):
    # def plot_bounding_box(ax, bbox, facecolor='none'):
    """
    Plots a bounding box on the given axis.

    Parameters:
    - ax: The matplotlib axis to plot on.
    - bbox: A tuple or list in the form (x_min, x_max, y_min, y_max).
    - edgecolor: The color of the bounding box edge.
    - facecolor: The fill color of the bounding box (default is transparent).
    """
    x_min, x_max, y_min, y_max = bbox
    width = x_max - x_min
    height = y_max - y_min

    # Create a rectangle patch
    rect = patches.Rectangle((x_min, y_min), width, height, linewidth=1, edgecolor=edgecolor, facecolor=facecolor)

    # Add the rectangle to the plot
    ax.add_patch(rect)


def generate_random_circle_multiple_exclusions_with_refPt(bounds, no_fly_zones, reference_point=None, min_distance=20):
    xmin, xmax, ymin, ymax = bounds

    # Start with the entire space as possible regions
    possible_regions = [(xmin, xmax, ymin, ymax)]

    for no_fly_bounds in no_fly_zones:
        no_fly_xmin, no_fly_xmax, no_fly_ymin, no_fly_ymax = no_fly_bounds
        new_regions = []

        for region in possible_regions:
            rxmin, rxmax, rymin, rymax = region

            # Check if the no-fly zone intersects with the current region
            if rxmin < no_fly_xmax and rxmax > no_fly_xmin and rymin < no_fly_ymax and rymax > no_fly_ymin:
                # Split the region into parts that are outside the no-fly zone
                if rxmin < no_fly_xmin:
                    new_regions.append((rxmin, no_fly_xmin, rymin, rymax))  # Left part
                if rxmax > no_fly_xmax:
                    new_regions.append((no_fly_xmax, rxmax, rymin, rymax))  # Right part
                if rymin < no_fly_ymin:
                    new_regions.append(
                        (max(rxmin, no_fly_xmin), min(rxmax, no_fly_xmax), rymin, no_fly_ymin))  # Bottom part
                if rymax > no_fly_ymax:
                    new_regions.append(
                        (max(rxmin, no_fly_xmin), min(rxmax, no_fly_xmax), no_fly_ymax, rymax))  # Top part
            else:
                # If no intersection, keep the entire region
                new_regions.append(region)

        possible_regions = new_regions

    # If a reference point is provided, exclude regions within the min_distance
    if reference_point is not None:
        ref_x, ref_y = reference_point
        new_regions = []
        for region in possible_regions:
            rxmin, rxmax, rymin, rymax = region

            # Calculate the closest point on the region boundary to the reference point
            closest_x = np.clip(ref_x, rxmin, rxmax)
            closest_y = np.clip(ref_y, rymin, rymax)

            # Calculate the distance from the reference point to the closest point in the region
            distance = np.sqrt((ref_x - closest_x) ** 2 + (ref_y - closest_y) ** 2)

            if distance >= min_distance:
                new_regions.append(region)

        possible_regions = new_regions

    # Randomly select one of the remaining regions
    if not possible_regions:
        raise ValueError("No valid regions available outside the no-fly zones and exclusion region.")

    selected_region = possible_regions[np.random.randint(0, len(possible_regions))]

    # Generate a random center within the selected region
    center_x = np.random.uniform(selected_region[0], selected_region[1])
    center_y = np.random.uniform(selected_region[2], selected_region[3])

    return [center_x, center_y]


def generate_random_circle_multiple_exclusions(bounds, no_fly_zones):
    xmin, xmax, ymin, ymax = bounds

    # Start with the entire space as possible regions
    possible_regions = [(xmin, xmax, ymin, ymax)]

    for no_fly_bounds in no_fly_zones:
        no_fly_xmin, no_fly_xmax, no_fly_ymin, no_fly_ymax = no_fly_bounds
        new_regions = []

        for region in possible_regions:
            rxmin, rxmax, rymin, rymax = region

            # Check if the no-fly zone intersects with the current region
            if rxmin < no_fly_xmax and rxmax > no_fly_xmin and rymin < no_fly_ymax and rymax > no_fly_ymin:
                # Split the region into parts that are outside the no-fly zone
                if rxmin < no_fly_xmin:
                    new_regions.append((rxmin, no_fly_xmin, rymin, rymax))  # Left part
                if rxmax > no_fly_xmax:
                    new_regions.append((no_fly_xmax, rxmax, rymin, rymax))  # Right part
                if rymin < no_fly_ymin:
                    new_regions.append(
                        (max(rxmin, no_fly_xmin), min(rxmax, no_fly_xmax), rymin, no_fly_ymin))  # Bottom part
                if rymax > no_fly_ymax:
                    new_regions.append(
                        (max(rxmin, no_fly_xmin), min(rxmax, no_fly_xmax), no_fly_ymax, rymax))  # Top part
            else:
                # If no intersection, keep the entire region
                new_regions.append(region)

        possible_regions = new_regions

    # Randomly select one of the remaining regions
    if not possible_regions:
        raise ValueError("No valid regions available outside the no-fly zones.")

    selected_region = possible_regions[np.random.randint(0, len(possible_regions))]

    # Generate a random center within the selected region
    center_x = np.random.uniform(selected_region[0], selected_region[1])
    center_y = np.random.uniform(selected_region[2], selected_region[3])

    # # Randomly select a radius that fits within the selected region
    # max_possible_radius_x = min(center_x - selected_region[0], selected_region[1] - center_x)
    # max_possible_radius_y = min(center_y - selected_region[2], selected_region[3] - center_y)
    # max_possible_radius = min(max_possible_radius_x, max_possible_radius_y, max_radius)
    #
    # radius = np.random.uniform(0, max_possible_radius)

    return [center_x, center_y]


def load_svg_image(svg_path):
    png_image_data = cairosvg.svg2png(url=svg_path)
    image = Image.open(io.BytesIO(png_image_data))
    return image


def polygons_single_cloud_conflict(circle, cloud_polygon):
    conflicts = []
    if not cloud_polygon.touches(circle) and cloud_polygon.intersects(circle):
        conflicts.append(cloud_polygon)
    elif cloud_polygon.within(circle):
        conflicts.append(cloud_polygon)
    return conflicts


def calculate_next_position(start_pos, target_pos, speed, time_step):
    # Calculate the direction vector from start to target
    direction_vector = target_pos - start_pos

    # Normalize the direction vector to get the unit direction vector
    distance_to_target = np.linalg.norm(direction_vector)
    if distance_to_target < 1:
        unit_direction_vector = np.zeros(
            2)  # prevent the case where current pos is very near to end pos which leads divide by zero
    else:
        unit_direction_vector = direction_vector / distance_to_target

    # Calculate the distance the agent will travel in one time step
    distance_travelled = speed * time_step

    # Calculate the new position
    new_position = start_pos + unit_direction_vector * distance_travelled

    return new_position


def calculate_bearing(x_host, y_host, x_intruder, y_intruder):
    delta_x = x_intruder - x_host
    delta_y = y_intruder - y_host

    theta_radians = math.atan2(delta_y, delta_x)
    theta_degrees = math.degrees(theta_radians)

    # Convert to bearing as specified
    if theta_degrees < 0:
        bearing = -theta_degrees
    else:
        bearing = 360 - theta_degrees

    return bearing


def initialize_excel_file(file_path):
    # Create a new workbook and add three empty sheets
    wb = Workbook()
    wb.save(file_path)


def append_to_excel(file_path, data):
    # Try to load the workbook, and create it if it does not exist
    try:
        wb = load_workbook(file_path)
        new_workbook = False
    except FileNotFoundError:
        wb = Workbook()
        new_workbook = True
        # wb.remove(wb.active)  # Remove the default sheet created by openpyxl if not need
    except Exception as e:  # Catch other exceptions, such as invalid file format
        print(f"An error occurred: {e}")
        return
    if new_workbook:
        if 'Sheet' in wb.sheetnames:
            # If the sheet exists, remove it
            wb.remove('Sheet')
    # Check if the required sheets exist, if not create them
    if len(data) == 3:  # for record eps time
        number_of_sheets = len(data)
        required_sheets = ['Sheet' + str(i) for i in range(number_of_sheets)]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(title=sheet_name)

        # Append the data to the respective sheets
        for i, portion in enumerate(data):
            sheet = wb[required_sheets[i]]
            if isinstance(portion, list):
                for item in portion:
                    sheet.append(item if isinstance(item, list) else [item])
                # If this is the last portion of data, append -1 to the last row
                if i == len(data) - 1:
                    sheet.append([-1])
            else:
                sheet.append([portion])
    else:
        if isinstance(data[0], list):  # for record reward
            # In reward record, we are recording in the form that one step have 3 agents
            required_sheets = ["agent0", "agent1", "agent2"]
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(title=sheet_name)
            for idx, step_reward in enumerate(data):
                for agent_idx, single_sheet_name in enumerate(required_sheets):
                    sheet = wb[single_sheet_name]
                    sheet.append(step_reward[agent_idx])
                    if idx == len(data) - 1:
                        sheet.append([-9999])

        else:  # for record noise
            wb.create_sheet(title='noise given at each step')
            sheet = wb['noise given at each step']
            for i, portion in enumerate(data):
                sheet.append(list(portion))
            sheet.append([-999])

    # Save the workbook
    wb.save(file_path)


def animate(frame_num, ax, env, trajectory_eachPlay):
    ax.clear()
    plt.axis('equal')
    plt.xlim(env.bound[0], env.bound[1])
    plt.ylim(env.bound[2], env.bound[3])
    plt.axvline(x=env.bound[0], c="green")
    plt.axvline(x=env.bound[1], c="green")
    plt.axhline(y=env.bound[2], c="green")
    plt.axhline(y=env.bound[3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")
    aircraft_svg_path = r'C:\Users\aiden.pang\Aircraft.svg'  # Replace with your SVG path
    plane_img = load_svg_image(aircraft_svg_path)
    # Define colors with transparency (alpha)
    colors = [
        (0.5, 0, 0.5),  # Purple
        (0.2, 0.8, 0.2),  # Lime
        (1, 0, 0),  # Red
        (0, 1, 0),  # Green
        (0, 0, 1),  # Blue
        (0, 1, 1),  # Cyan
        (1, 0, 1),  # Magenta
        (1, 1, 0),  # Yellow
        (1, 0.65, 0),  # Orange
    ]

    for line_idx, line in enumerate(env.potential_ref_line):
        x, y = line.xy
        plt.plot(x, y, linestyle='solid', linewidth=10, color=colors[line_idx], alpha=0.2)
        plt.plot(line.coords[0][0], line.coords[0][1], marker=MarkerStyle("^"), color=colors[line_idx])
        plt.plot(line.coords[-1][0], line.coords[-1][1], marker='*', color=colors[line_idx])

    # for agentIdx, agent in env.all_agents.items():
    #     plt.plot(agent.ini_pos[0], agent.ini_pos[1],
    #              marker=MarkerStyle("^"), color=colors[agentIdx])
    #     plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
    #     plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color=colors[agentIdx], markersize=10)
    #     plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)

        # # link individual drone's starting position with its goal
        # ini = agent.ini_pos
        # for wp in agent.ref_line.coords:
        #     plt.plot([wp[0], ini[0]], [wp[1], ini[1]], linestyle='solid', linewidth=10, color=colors[agentIdx],
        #              alpha=0.2)
        #     ini = wp

    # display cloud
    interval = 5  # Change cluster coordinates around centre every 10 frames
    for cloud_idx, cloud_agent in enumerate(env.cloud_config):
        # Define the fixed center
        center_x, center_y = cloud_agent.trajectory[frame_num].x, cloud_agent.trajectory[frame_num].y
        cloud_centre = Point(center_x, center_y)
        cloud_poly = cloud_centre.buffer(cloud_agent.radius)
        # ___add boundary circle for clouds---
        # matp_poly = shapelypoly_to_matpoly(cloud_poly, False, 'blue')  # the 3rd parameter is the edge color
        # matp_poly.set_zorder(5)
        # ax.add_patch(matp_poly)
        # Generate multiple clusters of random points within the specified range
        num_points_per_cluster = 5000
        num_clusters = 15
        x_range = cloud_agent.spawn_cluster_pt_x_range
        y_range = cloud_agent.spawn_cluster_pt_y_range
        if frame_num % interval == 0:
            cluster_centers_x = np.random.uniform(center_x + x_range[0], center_x + x_range[1], num_clusters)
            cluster_centers_y = np.random.uniform(center_y + y_range[0], center_y + y_range[1], num_clusters)
            cluster_centers = np.column_stack((cluster_centers_x, cluster_centers_y))
            cloud_agent.cluster_centres = cluster_centers

        # Generate points for each cluster with controlled density
        x, y = [], []
        for cx, cy in cloud_agent.cluster_centres:
            angles = np.random.uniform(0, 2 * np.pi, num_points_per_cluster)
            radii = np.random.normal(0, 0.1, num_points_per_cluster)  # Decrease spread for higher density
            x.extend(cx + radii * np.cos(angles))
            y.extend(cy + radii * np.sin(angles))
        x = np.array(x)
        y = np.array(y)
        # Create a 2D histogram to serve as the contour data
        margin = 25
        contour_min_x = center_x + x_range[0] - margin
        contour_max_x = center_x + x_range[1] + margin
        contour_min_y = center_y + y_range[0] - margin
        contour_max_y = center_y + y_range[1] + margin
        hist, xedges, yedges = np.histogram2d(x, y, bins=(100, 100),
                                              range=[[contour_min_x, contour_max_x], [contour_min_y, contour_max_y]])
        # Smooth the histogram to create a more organic shape
        hist = gaussian_filter(hist, sigma=5)  # Adjust sigma for better control

        # Create the custom colormap from green to yellow to red
        cmap = LinearSegmentedColormap.from_list('green_yellow_red', ['green', 'yellow', 'red'])
        X, Y = np.meshgrid(xedges[:-1] + 0.5 * (xedges[1] - xedges[0]), yedges[:-1] + 0.5 * (yedges[1] - yedges[0]))
        contour_levels = np.linspace(hist.min(), hist.max(), 10)
        contour = ax.contourf(X, Y, hist, levels=contour_levels, cmap=cmap)

        level_color = cmap(
            (contour_levels[1] - contour_levels.min()) / (contour_levels.max() - contour_levels.min()))
        # Extract the outermost contour path and overlay it with a black line
        outermost_contour = ax.contour(X, Y, hist, levels=[contour_levels[1]], colors=[level_color],
                                       linewidths=1)  # this line must be present to
        # Extract the vertices of the outermost contour path
        outermost_path = outermost_contour.collections[0].get_paths()[0]
        vertices = outermost_path.vertices
        x_clip, y_clip = vertices[:, 0], vertices[:, 1]
        # ax.plot(x_clip, y_clip, color="crimson")
        coordinates = np.column_stack((x_clip, y_clip))
        clippath = Path(coordinates)
        patch = PathPatch(clippath, facecolor='none')
        ax.add_patch(patch)
        for c in contour.collections:
            c.set_clip_path(patch)

    for a_idx, agent in enumerate(trajectory_eachPlay[frame_num]):
        x, y = agent[0], agent[1]

        # draw agent prob line
        for line_angle, line_info in agent[4].items():
            line_to_draw = LineString([Point(x, y), line_info[0]])
            line_x, line_y = line_to_draw.xy
            # plt.text(line_info[0].x, line_info[0].y, str(line_angle))
            ax.plot(line_x, line_y, color='k', linewidth=1)

        # plt.text(x-1, y-1, 'agent_'+str(a_idx)+'_'+str(round(float(frame_num), 2)))
        if np.issubdtype(agent[2].dtype, np.integer):
            # plt.text(x - 3, y - 3, 'a_' + str(a_idx) + '_' + str(agent[2]))
            plt.text(x - 3, y - 3, 'a' + str(a_idx))
        else:
            # plt.text(x - 3, y - 3, 'a_' + str(a_idx) + '_' + str(np.round((agent[2]), 4)))
            plt.text(x - 3, y - 3, 'a' + str(a_idx))
        heading = agent[3] * 180 / np.pi  # in degree
        img_extent = [
            x - env.all_agents[0].protectiveBound,
            x + env.all_agents[0].protectiveBound,
            y - env.all_agents[0].protectiveBound,
            y + env.all_agents[0].protectiveBound
        ]
        transform = Affine2D().rotate_deg_around(x, y, heading - 90) + ax.transData
        ax.imshow(plane_img, extent=img_extent, zorder=10, transform=transform)
        self_circle = Point(x, y).buffer(env.all_agents[0].protectiveBound, cap_style='round')
        grid_mat_Scir = shapelypoly_to_matpoly(self_circle, inFill=True, Edgecolor=None,
                                               FcColor='lightblue')  # None meaning no edge
        grid_mat_Scir.set_zorder(2)
        grid_mat_Scir.set_alpha(0.9)  # Set transparency to 0.5
        ax.add_patch(grid_mat_Scir)

    return ax.patches + [ax.texts]


def get_history_tensor(history, sequence_length, input_size):
    # If history is shorter than the sequence length, pad it with zeros
    if len(history) < sequence_length:
        padding = torch.zeros(sequence_length - len(history), input_size)
        history_tensor = torch.cat((padding, torch.stack(list(history))))
    else:
        history_tensor = torch.stack(list(history))

    # Add a batch dimension and return
    return history_tensor.unsqueeze(0)


def save_gif(env, trajectory_eachPlay, pre_fix, episode_to_check, episode):
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)

    plt.axis('equal')
    plt.xlim(env.bound[0], env.bound[1])
    plt.ylim(env.bound[2], env.bound[3])
    plt.axvline(x=env.bound[0], c="black")
    plt.axvline(x=env.bound[1], c="black")
    plt.axhline(y=env.bound[2], c="black")
    plt.axhline(y=env.bound[3], c="black")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")

    # draw occupied_poly
    for one_poly in env.world_map_2D_polyList[0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
    #     ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in env.world_map_2D_polyList[0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        # ax.add_patch(matp_poly)

    for agentIdx, agent in env.all_agents.items():
        plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                 marker=MarkerStyle(">",
                                    fillstyle="right",
                                    transform=Affine2D().rotate_deg(math.degrees(agent.heading))),
                 color='y')
        plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
        # plot self_circle of the drone
        self_circle = Point(agent.ini_pos[0],
                            agent.ini_pos[1]).buffer(agent.protectiveBound, cap_style='round')
        grid_mat_Scir = shapelypoly_to_matpoly(self_circle, inFill=False, Edgecolor='k')
        ax.add_patch(grid_mat_Scir)

        # plot drone's detection range
        detec_circle = Point(agent.ini_pos[0],
                             agent.ini_pos[1]).buffer(agent.detectionRange / 2, cap_style='round')
        detec_circle_mat = shapelypoly_to_matpoly(detec_circle, inFill=False, Edgecolor='g')
        ax.add_patch(detec_circle_mat)

        plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='y', markersize=10)
        plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)

    # Create animation
    ani = animation.FuncAnimation(fig, animate, fargs=(ax, env, trajectory_eachPlay), frames=len(trajectory_eachPlay),
                                  interval=300, blit=False)
    # Save as GIF
    gif_path = pre_fix + '\episode_' + str(episode_to_check) + 'simulation_num_' + str(episode) + '.gif'
    ani.save(gif_path, writer='pillow')

    # Close figure
    plt.close(fig)


def view_static_traj(env, trajectory_eachPlay, save_path=None, max_time_step=None):
    aircraft_svg_path = r'F:\githubClone\HotspotResolver_24\pictures\Aircraft.svg'  # Replace with your SVG path
    plane_img = load_svg_image(aircraft_svg_path)
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    colors = [
        (0.5, 0, 0.5),  # Purple
        (0.2, 0.8, 0.2),  # Lime
        (1, 0, 0),  # Red
        (0, 1, 0),  # Green
        (0, 0, 1),  # Blue
        (0, 1, 1),  # Cyan
        (1, 0, 1),  # Magenta
        (1, 1, 0),  # Yellow
        (1, 0.65, 0),  # Orange
    ]
    # display initial condition
    FixedAR_names = ['N884', 'M768', 'M767']
    for line_idx, line in enumerate(env.potential_ref_line):
        x, y = line.xy
        plt.plot(x, y, linestyle='solid', linewidth=10, color=colors[line_idx], alpha=0.2)
        plt.plot(line.coords[0][0], line.coords[0][1], marker=MarkerStyle("^"), color=colors[line_idx])  # start point
        plt.text(line.coords[0][0]+5, line.coords[0][1]+5, FixedAR_names[line_idx])
        plt.plot(line.coords[-1][0], line.coords[-1][1], marker='*', color=colors[line_idx])  # end point

    # for agentIdx, agent in env.all_agents.items():
    #     x, y = agent.pos[0], agent.pos[1]
    #     heading = agent.heading * 180 / np.pi  # in degree
    #     img_extent = [
    #         x - env.all_agents[0].protectiveBound,
    #         x + env.all_agents[0].protectiveBound,
    #         y - env.all_agents[0].protectiveBound,
    #         y + env.all_agents[0].protectiveBound
    #     ]
    #     transform = Affine2D().rotate_deg_around(x, y, heading - 90) + ax.transData
    #     ax.imshow(plane_img, extent=img_extent, zorder=10, transform=transform)
    #     plt.plot(agent.ini_pos[0], agent.ini_pos[1],
    #              marker=MarkerStyle("^"), color=colors[agentIdx])
    #     # plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
    #     # plot self_circle of the drone
    #     self_circle = Point(x, y).buffer(agent.protectiveBound, cap_style='round')
    #     grid_mat_Scir = shapelypoly_to_matpoly(self_circle, inFill=True, Edgecolor=None,
    #                                            FcColor='lightblue')  # None meaning no edge
    #     grid_mat_Scir.set_zorder(2)
    #     grid_mat_Scir.set_alpha(0.9)  # Set transparency to 0.5
    #     ax.add_patch(grid_mat_Scir)

    # draw trajectory in current episode
    if max_time_step is None:
        max_time_step = len(trajectory_eachPlay)

    # display cloud
    interval = 10  # Change cluster coordinates around centre every 10 frames
    # Calculate alpha values that will create a fading effect
    alpha_values = np.linspace(0.1, 1.0, max_time_step)
    contour_drawn = [False] * len(env.cloud_config)
    outline_drawn = [False] * len(env.cloud_config)


    for cloud_idx, cloud_agent in enumerate(env.cloud_config):
        for trajectory_idx in range(max_time_step):
            if trajectory_idx >= max_time_step:
                break
            if trajectory_idx % interval == 0:
                # Define the fixed center
                if max_time_step == len(trajectory_eachPlay):
                    center_x, center_y = cloud_agent.trajectory[trajectory_idx].x, cloud_agent.trajectory[trajectory_idx].y
                else:
                    center_x, center_y = cloud_agent.trajectory[max_time_step].x, cloud_agent.trajectory[
                        max_time_step].y
                # ___add boundary circle for clouds---
                # cloud_centre = Point(center_x, center_y)
                # cloud_poly = cloud_centre.buffer(cloud_agent.radius)
                # matp_poly = shapelypoly_to_matpoly(cloud_poly, False, 'blue')  # the 3rd parameter is the edge color
                # matp_poly.set_zorder(5)
                # ax.add_patch(matp_poly)
                # Generate multiple clusters of random points within the specified range
                num_points_per_cluster = 5000
                num_clusters = 15
                x_range = cloud_agent.spawn_cluster_pt_x_range
                y_range = cloud_agent.spawn_cluster_pt_y_range

                cluster_centers_x = np.random.uniform(center_x + x_range[0], center_x + x_range[1], num_clusters)
                cluster_centers_y = np.random.uniform(center_y + y_range[0], center_y + y_range[1], num_clusters)
                cluster_centers = np.column_stack((cluster_centers_x, cluster_centers_y))
                cloud_agent.cluster_centres = cluster_centers

                # Generate points for each cluster with controlled density
                x, y = [], []
                for cx, cy in cloud_agent.cluster_centres:
                    angles = np.random.uniform(0, 2 * np.pi, num_points_per_cluster)
                    radii = np.random.normal(0, 0.1, num_points_per_cluster)  # Decrease spread for higher density
                    x.extend(cx + radii * np.cos(angles))
                    y.extend(cy + radii * np.sin(angles))
                x = np.array(x)
                y = np.array(y)
                # Create a 2D histogram to serve as the contour data
                margin = 25
                contour_min_x = center_x + x_range[0] - margin
                contour_max_x = center_x + x_range[1] + margin
                contour_min_y = center_y + y_range[0] - margin
                contour_max_y = center_y + y_range[1] + margin
                hist, xedges, yedges = np.histogram2d(x, y, bins=(100, 100),
                                                      range=[[contour_min_x, contour_max_x], [contour_min_y, contour_max_y]])
                # Smooth the histogram to create a more organic shape
                hist = gaussian_filter(hist, sigma=5)  # Adjust sigma for better control

                # Create the custom colormap from green to yellow to red
                cmap = LinearSegmentedColormap.from_list('green_yellow_red', ['green', 'yellow', 'red'])
                X, Y = np.meshgrid(xedges[:-1] + 0.5 * (xedges[1] - xedges[0]), yedges[:-1] + 0.5 * (yedges[1] - yedges[0]))
                contour_levels = np.linspace(hist.min(), hist.max(), 10)

                if max_time_step == len(trajectory_eachPlay):
                    contour = ax.contourf(X, Y, hist, levels=contour_levels, cmap=cmap, alpha=alpha_values[trajectory_idx])
                else:
                    # in this loop we only required to draw once.
                    if contour_drawn[cloud_idx] == False:
                        contour = ax.contourf(X, Y, hist, levels=contour_levels, cmap=cmap)
                        contour_drawn[cloud_idx] = True
                level_color = cmap(
                    (contour_levels[1] - contour_levels.min()) / (contour_levels.max() - contour_levels.min()))
                # Extract the outermost contour path and overlay it with a black line

                if max_time_step == len(trajectory_eachPlay):
                    outermost_contour = ax.contour(X, Y, hist, levels=[contour_levels[1]], colors=[level_color],
                                                   linewidths=1, alpha=alpha_values[trajectory_idx])  # this line must be present to show the boundary fading
                else:
                    # in this loop we only required to draw once.
                    if outline_drawn[cloud_idx] == False:
                        outermost_contour = ax.contour(X, Y, hist, levels=[contour_levels[1]], colors=[level_color],
                                                       linewidths=1)
                        outline_drawn[cloud_idx] = True
                # Extract the vertices of the outermost contour path
                outermost_path = outermost_contour.collections[0].get_paths()[0]
                vertices = outermost_path.vertices
                x_clip, y_clip = vertices[:, 0], vertices[:, 1]
                # ax.plot(x_clip, y_clip, color="crimson")
                coordinates = np.column_stack((x_clip, y_clip))
                clippath = Path(coordinates)
                patch = PathPatch(clippath, facecolor='none', alpha=alpha_values[trajectory_idx])
                ax.add_patch(patch)
                for c in contour.collections:
                    c.set_clip_path(patch)

    for agentIDX, agent in env.all_agents.items():
        previous_position = agent.ini_pos  # Start with the agent's initial position
        for trajectory_idx in range(max_time_step):
            if trajectory_idx >= max_time_step:
                break
            each_agent_traj = trajectory_eachPlay[trajectory_idx][agentIDX]
            x, y = each_agent_traj[0], each_agent_traj[1]

            # Draw the trajectory as dotted lines starting from the initial position
            if trajectory_idx > 0:  # Ensure we're not drawing a redundant line from ini_pos to itself
                plt.plot([previous_position[0], x], [previous_position[1], y], linestyle=(0, (1, 10)),
                         color=colors[agentIDX])
            # Update previous position
            previous_position = (x, y)

            if trajectory_idx == max_time_step - 1:
                # Final position with aircraft marker
                heading = each_agent_traj[3] * 180 / np.pi  # Convert to degrees
                img_extent = [
                    x - env.all_agents[0].protectiveBound,
                    x + env.all_agents[0].protectiveBound,
                    y - env.all_agents[0].protectiveBound,
                    y + env.all_agents[0].protectiveBound
                ]
                transform = Affine2D().rotate_deg_around(x, y, heading - 90) + ax.transData
                if agent.ini_eta == None or agent.ini_eta < trajectory_idx:
                    ax.imshow(plane_img, extent=img_extent, zorder=10, transform=transform)

                # Draw the protective boundary around the final position
                self_circle = Point(x, y).buffer(env.all_agents[0].protectiveBound, cap_style='round')
                grid_mat_SCir = shapelypoly_to_matpoly(self_circle, inFill=True, Edgecolor=None, FcColor='lightblue')
                grid_mat_SCir.set_zorder(2)
                grid_mat_SCir.set_alpha(0.9)
                if agent.ini_eta == None or agent.ini_eta < trajectory_idx:
                    ax.add_patch(grid_mat_SCir)
                    plt.text(x+3, y+3, 'a'+str(agentIDX))

    # plt.axis('equal')
    plt.xlim(env.bound[0], env.bound[1])
    plt.ylim(env.bound[2], env.bound[3])
    # plt.axvline(x=env.bound[0], c="green")
    # plt.axvline(x=env.bound[1], c="green")
    # plt.axhline(y=env.bound[2], c="green")
    # plt.axhline(y=env.bound[3], c="green")
    plt.xlabel("Length (nm)")
    plt.ylabel("Width (nm)")

    # Save the figure if save_path is provided
    if save_path:
        # save svg
        svg_path = os.path.splitext(save_path)[0] + '.svg'
        plt.savefig(svg_path, bbox_inches='tight')
        # save pdf
        pdf_path = os.path.splitext(save_path)[0] + '.pdf'
        plt.savefig(pdf_path, bbox_inches='tight')
        # print(f"Figure saved at {save_path}")

    # plt.show()


def compute_t_cpa_d_cpa_potential_col(other_pos, host_pos, other_vel, host_vel, other_bound, host_bound,
                                      total_possible_conf):
    rel_dist_withNeg = -1 * (other_pos - host_pos)  # relative distance, host-intru
    rel_vel = other_vel - host_vel  # get relative velocity, host-intru
    rel_vel_norm_withSQ = np.square(np.linalg.norm(rel_vel))  # square of norm
    if rel_vel_norm_withSQ == 0:  # meaning this neigh with host drone relative vel = 0, same spd
        tcpa = -10
        # check possible collision manually
        time_to_check = 1  # Check for collision after t seconds
        new_nei_pos = other_pos + (other_vel * time_to_check)
        new_host_pos = host_pos + (host_vel * time_to_check)
        d_tcpa = np.linalg.norm(new_host_pos - new_nei_pos)
        if d_tcpa < (other_bound + host_bound):
            total_possible_conf = total_possible_conf + 1

    else:
        tcpa = np.dot(rel_dist_withNeg, rel_vel) / rel_vel_norm_withSQ
        d_tcpa = np.linalg.norm(((rel_dist_withNeg * -1) + (rel_vel * tcpa)))

    if (tcpa <= 1) and (tcpa >= 0) and (
            d_tcpa < (other_bound + host_bound)):
        total_possible_conf = total_possible_conf + 1
    return (tcpa, d_tcpa, total_possible_conf)


def compute_projected_velocity(vehicle_velocity, reference_path, vehicle_position):
    # Find the closest point on the path to the vehicle's current position
    closest_point = nearest_points(reference_path, vehicle_position)[0]
    # Check if the closest point is a turning point
    turning_points = [Point(p) for p in reference_path.coords[1:-1]]  # Exclude the first and last points
    # If closest_point is a turning point, use the outgoing path segment for tangent calculation
    if closest_point in turning_points:
        # Find the index of the turning point in the list of points
        turning_index = turning_points.index(closest_point)

        # The outgoing segment starts from the turning point
        segment_start = np.array(reference_path.coords[turning_index + 1])
        segment_end = np.array(reference_path.coords[turning_index + 2])

        # Calculate the tangent vector (unit vector) of the segment
        tangent_vector = (segment_end - segment_start) / np.linalg.norm(segment_end - segment_start)

        # Project the vehicle's velocity onto the tangent vector
        projected_velocity = np.dot(vehicle_velocity, tangent_vector) * tangent_vector
    else:
        # If it's not a turning point, just find the index of the closest segment
        distance_along_line = reference_path.project(closest_point)
        # Find the segment that contains this distance
        cumulative_distance = 0
        for i in range(len(reference_path.coords) - 1):
            segment_start = np.array(reference_path.coords[i])
            segment_end = np.array(reference_path.coords[i + 1])
            segment_length = LineString([segment_start, segment_end]).length
            cumulative_distance += segment_length
            if cumulative_distance > distance_along_line:
                break
        # Compute the tangent vector as before
        tangent_vector = (segment_end - segment_start) / np.linalg.norm(segment_end - segment_start)
        # Project the vehicle's velocity onto the tangent vector
        projected_velocity = np.dot(vehicle_velocity, tangent_vector) * tangent_vector
    return projected_velocity


def find_index_of_min_first_element(lst):
    # Initialize the index and the minimum value
    min_index = 0
    min_value = lst[0][0]

    # Loop through the list to find the min value and its index
    for i, sublist in enumerate(lst):
        if sublist[0] < min_value:
            min_value = sublist[0]
            min_index = i

    return min_index


def total_length_to_end_of_line(initial_point, linestring):
    """
    Calculate the total distance from an initial point to the nearest point on the line and
    from there to the end of the line.

    Parameters:
    initial_point (tuple): The initial point as (x, y).
    linestring (LineString): The LineString object.

    Returns:
    float: The total distance from the initial point to the end of the LineString.
    """
    # Create a Point object from the tuple
    point = Point(initial_point)

    # Find the nearest point on the line to the initial point
    nearest_point_on_line = linestring.interpolate(linestring.project(point))

    # Calculate the distance from the initial point to the nearest point on the line
    distance_to_line = point.distance(nearest_point_on_line)

    # Calculate the distance from the nearest point on the line to the end of the line
    projected_distance = linestring.project(nearest_point_on_line)
    distance_to_end_of_line = linestring.length - projected_distance

    # Sum the distances to get the total distance
    total_distance = distance_to_line + distance_to_end_of_line

    return total_distance


def total_length_to_end_of_line_without_cross(initial_point, linestring):
    """
    Calculate the total distance from an initial point to the nearest point on the line and
    from there to the end of the line.

    Parameters:
    initial_point (tuple): The initial point as (x, y).
    linestring (LineString): The LineString object.

    Returns:
    float: The total distance from the initial point to the end of the LineString.
    """
    # Create a Point object from the tuple
    point = Point(initial_point)

    # Find the nearest point on the line to the initial point
    nearest_point_on_line = linestring.interpolate(linestring.project(point))

    # Calculate the distance from the nearest point on the line to the end of the line
    projected_distance = linestring.project(nearest_point_on_line)
    distance_to_end_of_line = linestring.length - projected_distance

    return distance_to_end_of_line


def sort_polygons(polygons):  # this sorting is left to right, but bottom to top. so, 0th is below 2nd. [[2,3],[0,1]]
    boxes = [polygon.bounds for polygon in polygons]
    sorted_boxes = sorted(boxes, key=lambda box: (box[1], box[0]))
    sorted_polygons = [polygons[boxes.index(box)] for box in sorted_boxes]
    return sorted_polygons


# this is to remove the need for the package descrete
def shapelypoly_to_matpoly(ShapelyPolgon, inFill=False, Edgecolor=None, FcColor='blue'):
    xcoo, ycoo = ShapelyPolgon.exterior.coords.xy
    matPolyConverted = matPolygon(xy=list(zip(xcoo, ycoo)), fill=inFill, edgecolor=Edgecolor, facecolor=FcColor)
    return matPolyConverted


def extract_individual_obs(combine_state, agent_idx):
    individual_obs = []
    self_obs = combine_state[agent_idx][0]
    self_obs_grid = combine_state[agent_idx][1]
    self_surround = combine_state[agent_idx][2]
    individual_obs = [self_obs, self_obs_grid, self_surround]
    # return individual_obs  # for old V5
    return combine_state[agent_idx]


def map_range(value, coe_a):
    # Calculate the normalized value
    out_min = -coe_a
    out_max = coe_a
    # tanh() has range of -1 to 1
    in_min = -1
    in_max = 1
    normalized = (value - in_min) / (in_max - in_min)
    # Map the normalized value to the output range
    mapped = out_min + (normalized * (out_max - out_min))
    # Return the mapped value
    return mapped


def compute_potential_conflict(cur_drone_pos, cur_drone_vel, cur_drone_protRad, cur_neigh_pos, cur_neigh_vel,
                               cur_neigh_protRad, cur_neigh_idx, current_ts):
    minus_rel_dist_before = -1 * (cur_drone_pos - cur_neigh_pos)  # always current drone - neighbours
    rel_vel_before = (cur_drone_vel - cur_neigh_vel)
    if not (np.any(cur_drone_vel) or np.any(cur_neigh_vel)):
        # print("All elements in both arrays are zero")
        rel_vel_before = rel_vel_before + 1e-10  # add a very small number to suppress division by zero warning

    rel_vel_SQnorm_before = np.square(np.linalg.norm(rel_vel_before))

    t_cpa_before = np.dot(minus_rel_dist_before, rel_vel_before) / rel_vel_SQnorm_before
    d_cpa_before = np.linalg.norm(((cur_drone_pos - cur_neigh_pos) + (rel_vel_before * t_cpa_before)))
    # time to potential conflict, t_cpa is the actual time in sec, so, if t_cap=3, meaning in order to reach d_cap, the decision making agent need to take 6 steps. Currently our
    # if (t_cpa_before >= 0) and (t_cpa_before <= 1.5) and (d_cpa_before < (cur_drone_protRad + cur_neigh_protRad)):
    #     # pc_list.append(cur_neigh_idx)
    #     pc_list[cur_neigh_idx] = [t_cpa_before, d_cpa_before]
    # double check
    # host_future_pos = cur_drone_pos + (t_cpa_before*cur_drone_vel)
    # intru_future_pos = cur_neigh_pos + (t_cpa_before*cur_neigh_vel)
    # doubleCheck_dcpa = np.linalg.norm((host_future_pos-intru_future_pos))
    if (t_cpa_before >= 0) and (t_cpa_before < 3):
        return [cur_neigh_idx, t_cpa_before, d_cpa_before]
    else:
        return []


def padding_list(max_grid_obs_dim, input_list):
    # padding actions
    if isinstance(input_list, np.float64):
        print("check")
    tobePad_gridObs = list(np.zeros(max_grid_obs_dim - len(input_list), dtype=int))
    padded_gridObs = input_list + tobePad_gridObs
    return padded_gridObs


def preprocess_batch_for_critic_net(input_state, batch_size):
    critic_own_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents
    critic_grid_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents
    critic_neigh_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents
    for batch_idx in range(batch_size):
        critic_own_cur_state = []
        critic_own_grid_state = []
        critic_own_neigh_state = []
        for agent_cur in input_state:
            critic_own_cur_state.append(agent_cur[0][batch_idx, :])
            critic_own_grid_state.append(agent_cur[1][batch_idx, :])
            # for neigh, first: max_nei_num X single_nei_features flatten to 1D array
            flat_nei = agent_cur[2][batch_idx, :].flatten()  # default is flatten in a row.
            critic_own_neigh_state.append(flat_nei)
        critic_own_batched_cur_state.append(np.array(critic_own_cur_state).reshape((1, -1)))
        critic_grid_batched_cur_state.append(np.array(critic_own_grid_state).reshape((1, -1)))
        critic_neigh_batched_cur_state.append(np.array(critic_own_neigh_state).reshape((1, -1)))

    cur_state_pre_processed = [T.tensor(np.array(critic_own_batched_cur_state)),
                               T.tensor(np.array(critic_grid_batched_cur_state)),
                               T.tensor(np.array(critic_neigh_batched_cur_state))]
    return cur_state_pre_processed


def preprocess_batch_for_critic_net_v2(input_state, batch_size):
    critic_own_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents

    for batch_idx in range(batch_size):
        critic_own_cur_state = []

        for agent_cur in input_state:
            critic_own_cur_state.append(agent_cur[batch_idx, :])

        critic_own_batched_cur_state.append(np.array(critic_own_cur_state).reshape((1, -1)))

    cur_state_pre_processed = T.tensor(np.array(critic_own_batched_cur_state))  # batch X (1 x no_agent x feature size)

    return cur_state_pre_processed


class OrnsteinUhlenbeckProcess:

    def __init__(self, size, mu=0, theta=0.2, sigma=0.2):
        self.size = size
        self.mu = mu
        self.theta = theta
        self.sigma = sigma
        self.state = np.ones(self.size) * self.mu
        self.reset_states()

    def reset_states(self):
        self.state = np.ones(self.size) * self.mu

    def sample(self, changing_sigma):
        x = self.state
        dx = self.theta * (self.mu - x) + changing_sigma * np.random.randn(len(x))
        self.state = x + dx
        return self.state


class OUActionNoise:
    def __init__(self, mu, sigma, theta, dt):
        self.ou_gen = None
        self.sigma = sigma
        self.mu = mu
        self.theta = theta
        self.dt = dt

    def reset(self):
        self.ou_gen = ornstein_uhlenbeck_gen(self.mu, self.sigma, self.theta, self.dt)

    def __call__(self):
        return next(self.ou_gen)


def action_rescaler(low, high):
    halfspan = (high - low) / 2.0
    middle = (high + low) / 2.0

    def rescaler(action):
        return np.clip((action + middle) * halfspan, low, high)

    return rescaler


def ornstein_uhlenbeck_gen(
        mu: np.ndarray = None,
        sigma: float = 1.0,
        theta: float = 0.15,
        dt: float = 1e-2,
):
    rescaler = action_rescaler(-1, 1)  # this acts like an noise clipping, align with the paper's description
    noise_gen = ornstein_uhlenbeck_unscaled(mu=mu, sigma=sigma, theta=theta, dt=dt)
    for action in noise_gen:
        yield rescaler(action)


def ornstein_uhlenbeck_unscaled(
        mu: np.ndarray = None,
        sigma: float = 1.0,
        dim: int = None,
        theta: float = 0.15,
        dt: float = 1e-2,
):
    if mu is None:
        if not dim:
            raise ValueError("Either mu or dimensionality parameter dim must be set!")
        mu = np.zeros(dim)
    last_noise = np.zeros_like(mu)
    while True:
        noise = (
                last_noise
                + theta * (mu - last_noise) * dt
                + sigma * np.sqrt(dt) * np.random.normal(size=mu.shape)
        )
        last_noise = noise
        yield noise


class NormalizeData:
    def __init__(self, x_min_max, y_min_max, spd_max, acc_range):
        self.normalize_max = 1
        self.normalize_min = -1
        self.dis_min_x = x_min_max[0]
        self.dis_max_x = x_min_max[1]
        self.dis_min_y = y_min_max[0]
        self.dis_max_y = y_min_max[1]
        self.spd_max = spd_max
        self.acc_min = acc_range[0]
        self.acc_max = acc_range[1]
        self.scale_attribute()

    def scale_attribute(self):
        self.x_scale = (self.normalize_max - self.normalize_min) / (self.dis_max_x - self.dis_min_x)
        self.y_scale = (self.normalize_max - self.normalize_min) / (self.dis_max_y - self.dis_min_y)

    def nmlz_pos(self, pos_c):
        x, y = pos_c[0], pos_c[1]
        x_normalized = 2 * ((x - self.dis_min_x) / (self.dis_max_x - self.dis_min_x)) - 1
        y_normalized = 2 * ((y - self.dis_min_y) / (self.dis_max_y - self.dis_min_y)) - 1
        return np.array([x_normalized, y_normalized])

    def reverse_nmlz_pos(self, norm_pos_c):
        norm_x, norm_y = norm_pos_c[0], norm_pos_c[1]
        x = ((norm_x + 1) / 2) * (self.dis_max_x - self.dis_min_x) + self.dis_min_x
        y = ((norm_y + 1) / 2) * (self.dis_max_y - self.dis_min_y) + self.dis_min_y
        return np.array([x, y])

    def scale_pos(self,
                  pos_c):  # NOTE: this method is essentially same as min-max normalize approach, but we need this appraoch to calculate x & y scale
        x_normalized = self.normalize_min + (pos_c[0] - self.dis_min_x) * self.x_scale
        y_normalized = self.normalize_min + (pos_c[1] - self.dis_min_y) * self.y_scale
        return np.array([x_normalized, y_normalized])

    def norm_scale(self, change_in_pos):
        return np.array([self.x_scale * change_in_pos[0], self.y_scale * change_in_pos[1]])

    def nmlz_pos_diff(self, diff):
        dx, dy = diff[0], diff[1]
        dx_min = self.dis_min_x - self.dis_max_x
        dx_max = self.dis_max_x - self.dis_min_x
        dy_min = self.dis_min_y - self.dis_max_y
        dy_max = self.dis_max_y - self.dis_min_y
        dx_normalized = 2 * ((dx - dx_min) / (dx_max - dx_min)) - 1
        dy_normalized = 2 * ((dy - dy_min) / (dy_max - dy_min)) - 1
        return dx_normalized, dy_normalized

    def nmlz_vel(self, cur_vel):
        vx, vy = cur_vel[0], cur_vel[1]
        vx_normalized = vx / self.spd_max
        vy_normalized = vy / self.spd_max
        # vx_normalized = (vx / self.spd_max) * 2 - 1
        # vy_normalized = (vy / self.spd_max) * 2 - 1
        return np.array([vx_normalized, vy_normalized])

    def reverse_nmlz_vel(self, norm_vel):
        norm_vx, norm_vy = norm_vel[0], norm_vel[1]
        vx = norm_vx * self.spd_max
        vy = norm_vy * self.spd_max
        return np.array([vx, vy])

    def nmlz_acc(self, cur_acc):
        ax, ay = cur_acc[0], cur_acc[1]
        # ax_normalized = ((ax - self.acc_min) / (self.acc_max-self.acc_min)) * 2 - 1
        # ay_normalized = (ay - self.acc_min) / (self.acc_max-self.acc_min) * 2 - 1
        ax_normalized = ax / self.acc_max
        ay_normalized = ay / self.acc_max
        return np.array([ax_normalized, ay_normalized])


def BetaNoise(action, noise_scale):
    action = action.detach().numpy()  # since the input is a tensor we must convert it to numpy before operations
    sign = np.sign(action)  # tracking the sign so we can flip the samples later
    action = abs(action)  # we only use right tail of beta
    alpha = 1 / noise_scale  # this determines the how contentrated the beta dsn is
    value = 0.5 + action / 2  # converting from action space of -1 to 1 to beta space of 0 to 1
    beta = alpha * (1 - value) / value  # calculating beta
    beta = beta + 1.0 * (
            (alpha - beta) / alpha)  # adding a little bit to beta prevents actions getting stuck at -1 or 1
    sample = np.random.beta(alpha, beta)  # sampling from the beta distribution
    sample = sign * sample + (1 - sign) / 2  # flipping sample if sign is <0 since we only use right tail of beta dsn

    action_output = 2 * sample - 1  # converting back to action space -1 to 1
    return torch.tensor(action_output)  # converting back to tensor


def GaussNoise(action, noise_scale):
    n = np.random.normal(0, 1, len(action))  # create some standard normal noise
    return torch.clamp(action + torch.tensor(noise_scale * n).float(), -1, 1)  # add the noise to the actions


def WeightedNoise(action, noise_scale, action_type):
    """
    Returns the epsilon scaled noise distribution for adding to Actor
    calculated action policy.
    """
    if action_type == 'continuous':
        target = np.random.uniform(-1, 1, 2)  # the action space is -1 to 1
    elif action_type == 'discrete':
        target = np.random.uniform(0, 1, 4)  # action space is discrete
        target = target / sum(target)
    action = noise_scale * target + (
            1 - noise_scale) * action.detach().numpy()  # take a weighted average with noise_scale as the noise weight
    return torch.tensor(action).float()


def display_trajectory(cur_env, combined_trajectory, eps_to_watch):
    episode_to_show = eps_to_watch
    episode_steps = combined_trajectory[episode_to_show]
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    # # draw link towards destination for all drones, destination for each drone didn't change
    for agentIdx in cur_env.all_agents:
        plt.plot([cur_env.all_agents[agentIdx].ini_pos[0], cur_env.all_agents[agentIdx].goal[0][0]],
                 [cur_env.all_agents[agentIdx].ini_pos[1], cur_env.all_agents[agentIdx].goal[0][1]], '--', color='c')
    reward, done = [], []
    crash_penalty = -100
    reach_target = 1000
    potential_conflict_count = 0
    fixed_domino_reward = 1
    x_left_bound = LineString([(cur_env.bound[0], -9999), (cur_env.bound[0], 9999)])
    x_right_bound = LineString([(cur_env.bound[1], -9999), (cur_env.bound[1], 9999)])
    y_bottom_bound = LineString([(-9999, cur_env.bound[2]), (9999, cur_env.bound[2])])
    y_top_bound = LineString([(-9999, cur_env.bound[3]), (9999, cur_env.bound[3])])

    allBuildingSTR = STRtree(cur_env.world_map_2D_polyList[0][0])

    for step_idx, agents_traj in enumerate(episode_steps):
        step_R = []
        step_D = []
        for ea_idx, each_agent in enumerate(agents_traj):
            # propagate environment
            cur_env.all_agents[ea_idx].pre_pos = cur_env.all_agents[ea_idx].pos
            cur_env.all_agents[ea_idx].pos = np.array([each_agent[0], each_agent[1]])
            # check reward for each step
            # calculate the deviation from the reference path after an action has been taken
            curPoint = Point(cur_env.all_agents[ea_idx].pos)
            host_refline = LineString([cur_env.all_agents[ea_idx].ini_pos, cur_env.all_agents[ea_idx].goal[0]])
            cross_track_deviation = curPoint.distance(
                host_refline)  # deviation from the reference line, cross track error

            host_pass_line = LineString([cur_env.all_agents[ea_idx].pre_pos, cur_env.all_agents[ea_idx].pos])
            host_passed_volume = host_pass_line.buffer(cur_env.all_agents[ea_idx].protectiveBound, cap_style='round')
            possiblePoly = allBuildingSTR.query(host_passed_volume)
            for element in possiblePoly:
                if allBuildingSTR.geometries.take(element).intersection(host_passed_volume):
                    collide_building = 1
                    print("drone_{} crash into building when moving from {} to {} at time step {}".format(ea_idx,
                                                                                                          cur_env.all_agents[
                                                                                                              ea_idx].pre_pos,
                                                                                                          cur_env.all_agents[
                                                                                                              ea_idx].pos,
                                                                                                          step_idx))
                    break

            tar_circle = Point(cur_env.all_agents[ea_idx].goal[0]).buffer(1, cap_style='round')
            goal_cur_intru_intersect = host_passed_volume.intersection(tar_circle)
            if not goal_cur_intru_intersect.is_empty:  # reached goal?
                print("drone_{} has reached its goal at time step {}".format(ea_idx, step_idx))
                step_D.append(True)
                step_R.append(np.array(reach_target))
            # exceed bound condition, don't use current point, use current circle or else will have condition that
            elif x_left_bound.intersects(host_passed_volume) or x_right_bound.intersects(
                    host_passed_volume) or y_bottom_bound.intersects(host_passed_volume) or y_top_bound.intersects(
                    host_passed_volume):
                print("drone_{} has crash into boundary at time step {}".format(ea_idx, step_idx))
                step_R.append(np.array(crash_penalty))
                step_D.append(True)
            # exceed bound or crash into buildings or crash with other neighbors
            elif collide_building == 1 or len(collision_drones) > 0:
                reward.append(np.array(crash_penalty))
                done.append(True)
            else:  # a normal step taken
                step_D.append(False)
                crossCoefficient = 1
                goalCoefficient = 6
                # cross track error term
                cross_track_error = (20 / ((cross_track_deviation * cross_track_deviation) / 200 + 1)) - 3.5
                # Distance between drone and its goal for two consecutive time step
                before_dist_hg = np.linalg.norm(cur_env.all_agents[ea_idx].pre_pos - cur_env.all_agents[ea_idx].goal[0])
                after_dist_hg = np.linalg.norm(cur_env.all_agents[ea_idx].pos - cur_env.all_agents[ea_idx].goal[
                    0])  # distance to goal after action
                delta_hg = goalCoefficient * (before_dist_hg - after_dist_hg)
                # a small penalty for discourage the agent to stay in one single spot
                if (before_dist_hg - after_dist_hg) <= 2:
                    small_step_penalty = 50
                else:
                    small_step_penalty = 0
                # Domino term also use as an indicator for agent to avoid other drones. so no need to specifically
                # add a term to avoid surrounding drones
                # step_reward = crossCoefficient*cross_track_error + delta_hg + dominoTerm - small_step_penalty
                step_reward = crossCoefficient * cross_track_error + delta_hg - small_step_penalty
                # step_reward = delta_hg
                # convert to arr
                step_R.append(np.array(step_reward))
                plt.text(each_agent[0] + 5, each_agent[1], str(np.array(round(step_reward, 1))))

            # plot self_circle of the drone
            self_circle = Point(each_agent[0], each_agent[1]).buffer(2.5, cap_style='round')
            grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')

            # label drone time step for the position
            plt.text(each_agent[0], each_agent[1], str(ea_idx))
            # plt.text(each_agent[0]+5, each_agent[1], str(step_idx))

            ax.add_patch(grid_mat_Scir)
        reward.append(step_R)
        done.append(step_D)
    # print reward
    print_R = 0
    for eps_stepR in reward:
        print_R = print_R + sum(eps_stepR)
    print(print_R)

    # draw occupied_poly
    for one_poly in cur_env.world_map_2D_polyList[0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in cur_env.world_map_2D_polyList[0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in cur_env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    plt.axis('equal')
    plt.xlim(cur_env.bound[0], cur_env.bound[1])
    plt.ylim(cur_env.bound[2], cur_env.bound[3])
    plt.axvline(x=cur_env.bound[0], c="green")
    plt.axvline(x=cur_env.bound[1], c="green")
    plt.axhline(y=cur_env.bound[2], c="green")
    plt.axhline(y=cur_env.bound[3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")
    plt.show()
    return reward


def display_exploration_expolitation(cur_env, combined_trajectory, eps_period):
    episode_to_show = 4999
    episode_steps = combined_trajectory[episode_to_show]
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    selfLabel = 0
    # # draw link towards destination for all drones, destination for each drone didn't change
    for agentIdx in cur_env.all_agents:
        plt.plot([cur_env.all_agents[agentIdx].ini_pos[0], cur_env.all_agents[agentIdx].goal[0][0]],
                 [cur_env.all_agents[agentIdx].ini_pos[1], cur_env.all_agents[agentIdx].goal[0][1]], '--', color='c')
    x_explore = []
    y_explore = []
    x_exploit = []
    y_exploit = []
    for epsIDX, episode_steps in enumerate(combined_trajectory):
        for step_idx, agents_traj in enumerate(episode_steps):
            for ea_idx, each_agent in enumerate(agents_traj):
                if epsIDX <= eps_period:
                    x_explore.append(each_agent[0])
                    y_explore.append(each_agent[1])
                else:
                    x_exploit.append(each_agent[0])
                    y_exploit.append(each_agent[1])

                # plot self_circle of the drone
                if selfLabel == 0:
                    self_circle = Point(each_agent[0], each_agent[1]).buffer(2.5, cap_style='round')
                    grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')
                    ax.add_patch(grid_mat_Scir)

                    # check transition
                    # map output action from NN to actual range

                    coe_a = 4
                    timestep = 1
                    ax_, ay_ = 1, 1
                    ax_ = map_range(ax_, coe_a)
                    ay_ = map_range(ay_, coe_a)
                    curVelx = cur_env.all_agents[ea_idx].vel[0] + ax_ * timestep
                    curVely = cur_env.all_agents[ea_idx].vel[1] + ay_ * timestep
                    delta_x = curVelx * timestep
                    delta_y = curVely * timestep
                    cur_env.all_agents[ea_idx].pos = np.array([cur_env.all_agents[ea_idx].pos[0] + delta_x,
                                                               cur_env.all_agents[ea_idx].pos[1] + delta_y])
                    # plt.scatter(cur_env.all_agents[ea_idx].pos[0], cur_env.all_agents[ea_idx].pos[1], color='lightblue')

            selfLabel = 1

    # draw occupied_poly
    for one_poly in cur_env.world_map_2D_polyList[0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'w', 'g')  # 4th parameter is the face color
        # ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in cur_env.world_map_2D_polyList[0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'w')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in cur_env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    cmap_exploit = colors.LinearSegmentedColormap.from_list("", ["white", "yellow"])
    cmap_explore = colors.LinearSegmentedColormap.from_list("", ["white", "blue"])

    # Hexbin exploit(Yellow) with white color for zero count
    # hb1 = plt.hexbin(np.array(x_exploit), np.array(y_exploit), gridsize=25, cmap=cmap_exploit, mincnt=1, alpha=1)
    # cb1 = plt.colorbar(hb1)
    # cb1.set_label('Exploit Hexbin')

    # # Hexbin explore(blue) with white color for zero count
    hb2 = plt.hexbin(np.array(x_explore), np.array(y_explore), gridsize=50, cmap=cmap_explore, mincnt=1, alpha=1)
    cb2 = plt.colorbar(hb2)
    cb2.set_label('Explore Hexbin')

    plt.axis('equal')
    plt.xlim(cur_env.bound[0], cur_env.bound[1])
    plt.ylim(cur_env.bound[2], cur_env.bound[3])
    plt.axvline(x=cur_env.bound[0], c="green")
    plt.axvline(x=cur_env.bound[1], c="green")
    plt.axhline(y=cur_env.bound[2], c="green")
    plt.axhline(y=cur_env.bound[3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")
    plt.show()


def action_selection_statistics(action_selection_collection):
    all_action_collection_x = []
    all_action_collection_y = []
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    for each_eps_collection in action_selection_collection:
        for each_step in each_eps_collection:
            for agent_idx, agent_val in each_step.items():
                all_action_collection_x.append(agent_val[0][0])
                all_action_collection_y.append(agent_val[0][1])

    # Set the number of bins for x and y
    num_bins = 20
    # Create the 2D histogram
    plt.hist2d(all_action_collection_x, all_action_collection_y, bins=num_bins)

    # Set the x-axis and y-axis labels
    plt.xlabel('X')
    plt.ylabel('Y')

    # Set the title of the histogram
    plt.title('2D Histogram of X and Y')

    # Add a colorbar
    plt.colorbar()

    # Show the histogram
    plt.show()
