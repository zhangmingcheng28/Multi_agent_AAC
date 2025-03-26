# -*- coding: utf-8 -*-
"""
@Time    : 3/13/2023 1:28 PM
@Author  : Mingcheng
@FileName: 
@Description: 
@Package dependency:
"""
from shapely.geometry import MultiPoint
from matplotlib.patches import Polygon as matPolygon
import torch as T
import random
from matplotlib.patches import Rectangle, Circle
import matplotlib.patches as patches
import numpy as np
import torch
import os
import matplotlib
import matplotlib.pyplot as plt
from shapely.strtree import STRtree
from shapely.geometry import LineString, Point, Polygon
import matplotlib.colors as colors
from shapely.ops import nearest_points
from matplotlib.transforms import Affine2D
import matplotlib.animation as animation
from openpyxl import load_workbook
from openpyxl import Workbook
from matplotlib.markers import MarkerStyle
import math
from matplotlib.lines import Line2D
from matplotlib.legend_handler import HandlerPatch


# Custom handler for the square patch
class HandlerSquare(HandlerPatch):
    def __init__(self, square_size_factor):
        self.square_size_factor = square_size_factor
        super().__init__()

    def create_artists(self, legend, orig_handle, x0, y0, width, height, fontsize, trans):
        square_size = self.square_size_factor * min(width, height)
        square = plt.Rectangle((x0 + (width - square_size) / 2, y0 + (height - square_size) / 2), square_size, square_size, transform=trans, color='blue', alpha=0.9)
        return [square]


def generate_random_polygon(center, num_points=10, max_dim=10):
    """
    Randomly selects either the convex or nonconvex polygon generation function
    and returns the generated polygon.
    """
    polygon_func = random.choice([generate_convex_polygon, generate_nonconvex_polygon])
    return polygon_func(center, num_points, max_dim)


def create_corrected_small_circles_outside(center, exterior_coords, small_circle_radius, geo_fence_areas, occupied_grid_lines_STRtree):
    small_circles = []
    small_circles_coords = []
    for x, y in exterior_coords.coords:
        overlaps_or_within_or_cuts = False
        # Calculate the direction vector from the center of the big circle to the exterior point
        direction_vector = np.array([x - center.x, y - center.y]) / np.linalg.norm([x - center.x, y - center.y])

        # Translate the center of the small circle in the direction of the direction vector outside the big circle
        center_x, center_y = x + direction_vector[0] * small_circle_radius, y + direction_vector[
            1] * small_circle_radius
        small_circle = Point(center_x, center_y).buffer(small_circle_radius)

        # Check if the small circle overlaps with any of the big circles
        if not any(small_circle.overlaps(geo_fence) or small_circle.within(geo_fence) for geo_fence in geo_fence_areas):
            # check for overlaps or within the grid polygons
            potential_candiates_idx = occupied_grid_lines_STRtree.query(small_circle)
            for idx in potential_candiates_idx:
                element = occupied_grid_lines_STRtree.geometries[idx]
                if element.type == 'Polygon':
                    if small_circle.overlaps(element) or small_circle.within(element):
                        overlaps_or_within_or_cuts = True
                        break
                elif element.type == 'LineString':
                    if small_circle.intersects(element) and not small_circle.touches(element):
                        overlaps_or_within_or_cuts = True
                        break
            if not overlaps_or_within_or_cuts:
                small_circles.append(small_circle)
                small_circles_coords.append([small_circle.centroid.x, small_circle.centroid.y])
    return small_circles, small_circles_coords


def create_adjusted_circles_along_line(line, radius, interval, geo_fence_areas):
    circles = []
    circle_centre = []
    distances = np.arange(0, line.length, interval)
    for distance in distances:
        point = line.interpolate(distance)
        tem_circle = point.buffer(radius)
        if not any(tem_circle.overlaps(geo_fence) or tem_circle.within(geo_fence) for geo_fence in geo_fence_areas):
            circles.append(tem_circle)
            circle_centre.append([tem_circle.centroid.x, tem_circle.centroid.y])
    last_pt = line.interpolate(line.length)
    circles.append(last_pt.buffer(radius))  # Ensure the last circle is at the end
    circle_centre.append([last_pt.x, last_pt.y])
    return circles, circle_centre


def grayscale_to_rgba(value):
    return (value, value, value, 1)


def generate_convex_polygon(center, num_points=10, max_dim=10):
    """
    Generates a convex polygon around a designated center coordinate.

    Parameters:
        center (tuple): The designated center coordinate (x, y).
        num_points (int): Total number of points to generate (including the center).
        max_dim (float): Maximum dimension (side length) of the bounding square. Default is 10.

    Returns:
        shapely.geometry.Polygon: A convex polygon generated from the random points.
    """
    half_dim = max_dim / 2.0
    # Always include the center to ensure it is inside the polygon.
    points = [center]
    # Generate the remaining points uniformly inside the square centered at 'center'
    points.extend([
        (
            random.uniform(center[0] - half_dim, center[0] + half_dim),
            random.uniform(center[1] - half_dim, center[1] + half_dim)
        )
        for _ in range(num_points - 1)
    ])
    # Compute the convex hull of the points.
    convex_polygon = MultiPoint(points).convex_hull
    return convex_polygon


def generate_nonconvex_polygon(center, num_points=10, max_dim=10):
    """
    Generates a non-convex polygon around a designated center coordinate.

    This function starts by generating a convex polygon around the center, then
    perturbs its vertices to create indentations, resulting in a non-convex shape.

    Parameters:
        center (tuple): The designated center coordinate (x, y).
        num_points (int): Total number of points to generate for the initial convex polygon.
        max_dim (float): Maximum dimension (side length) of the bounding square. Default is 10.

    Returns:
        shapely.geometry.Polygon: A non-convex polygon.
    """
    # Generate a convex polygon around the center.
    convex_poly = generate_convex_polygon(center, num_points, max_dim)
    coords = list(convex_poly.exterior.coords)[:-1]  # Exclude the repeated last coordinate
    centroid = convex_poly.centroid

    new_coords = []
    for (x, y) in coords:
        # Compute the vector from the polygon's centroid to the vertex.
        vec = np.array([x - centroid.x, y - centroid.y])
        # Perturb the vertex inward or outward to create a concave effect.
        factor = random.uniform(0.5, 1.0)
        new_point = (centroid.x + vec[0] * factor, centroid.y + vec[1] * factor)
        new_coords.append(new_point)

    nonconvex_poly = Polygon(new_coords)
    # If the result is still convex, force a concave indentation.
    if nonconvex_poly.convex_hull.equals(nonconvex_poly):
        mid_x = (new_coords[0][0] + new_coords[1][0]) / 2
        mid_y = (new_coords[0][1] + new_coords[1][1]) / 2
        new_coords.insert(1, (mid_x, mid_y))
        nonconvex_poly = Polygon(new_coords)

    return nonconvex_poly


def spawn_obstacle(host_position, goal_position, occupancy_dict, spawn_distance=10, radius=5, candidate_points=None,
                   obstacles_list=None):
    """
    Finds a valid obstacle center that satisfies the following conditions:
      1. At least 'spawn_distance' away from the host UAV.
      2. The full circular area (radius) around the candidate is free.
      3. The candidate center is in front of the host UAV (toward goal_position).

    If found, updates the occupancy_dict by marking grid cells within the circle as occupied and
    optionally tracks the obstacle in obstacles_list.

    Parameters:
        host_position (tuple): (x, y) position of the host UAV.
        goal_position (tuple): (x, y) position of the host UAV's goal.
        occupancy_dict (dict): Occupancy grid with keys as (x, y, z) and values 0 (free) or 1 (occupied).
        spawn_distance (float): Minimum distance from the host UAV (default: 10 m).
        radius (float): Radius of the obstacle (default: 5 m).
        candidate_points (list, optional): Pre-computed list of candidate (x, y) points.
        obstacles_list (list, optional): List to track spawned obstacles.

    Returns:
        tuple or None: The valid obstacle center (x, y) if found and spawned; otherwise, None.
    """
    # Find a valid candidate center.
    valid_center = generate_valid_obstacle_center(host_position, goal_position, occupancy_dict, spawn_distance, radius,
                                                  candidate_points)
    if valid_center is None:
        print("No valid obstacle center found.")
        return None

    # Update the occupancy_dict: mark grid cells within the circle as occupied.
    for key in list(occupancy_dict.keys()):
        cell_x, cell_y = key
        if (cell_x - valid_center[0]) ** 2 + (cell_y - valid_center[1]) ** 2 <= radius ** 2:
            occupancy_dict[key] = 1

    # Optionally track the obstacle.
    if obstacles_list is not None:
        obstacles_list.append({'center': valid_center, 'radius': radius})

    print(f"Obstacle spawned at {valid_center} with radius {radius} m.")
    return valid_center


def generate_valid_obstacle_center(host_position, goal_position, occupancy_dict, spawn_distance=10, radius=5,
                                   candidate_points=None):
    """
    Searches for a valid obstacle center that satisfies:

    1. The center is at least 'spawn_distance' meters away from the host UAV.
    2. The full circular area (with given 'radius') around the center is free (all grid cells unoccupied).
    3. The candidate center is in front of the host UAV (in the direction of goal_position).

    Parameters:
        host_position (tuple): (x, y) position of the host UAV.
        goal_position (tuple): (x, y) position of the host UAV's goal.
        occupancy_dict (dict): Dictionary with keys as (x, y, z) and values 0 (free) or 1 (occupied).
        spawn_distance (float): Minimum required distance from host UAV (default: 10 m).
        radius (float): Radius of the circular area to check (default: 5 m).
        candidate_points (list, optional): A list of candidate (x, y) points. If None, this list is generated
                                           from occupancy_dict keys that are free.

    Returns:
        tuple or None: A valid (x, y) center if found; otherwise, None.
    """
    # Generate candidate points from occupancy_dict keys if not provided.
    if candidate_points is None:
        candidate_points = [(x, y) for (x, y) in occupancy_dict.keys() if occupancy_dict[(x, y)] == 0]

    # Filter candidates that are at least spawn_distance away from the host UAV.
    # valid_candidates = [
    #     pt for pt in candidate_points
    #     if math.hypot(pt[0] - host_position[0], pt[1] - host_position[1]) > spawn_distance
    # ]

    valid_candidates = [
        pt for pt in candidate_points
        if spawn_distance <= math.hypot(pt[0] - host_position[0], pt[1] - host_position[1]) <= spawn_distance+30
    ]

    # Shuffle to randomize the order.
    random.shuffle(valid_candidates)

    # Compute the vector from host to goal.
    dx = goal_position[0] - host_position[0]
    dy = goal_position[1] - host_position[1]
    norm_sq = dx**2 + dy**2

    for candidate in valid_candidates:
        # Create the vector from host to candidate.
        cx = candidate[0] - host_position[0]
        cy = candidate[1] - host_position[1]
        # Compute projection parameter t (0 means host, 1 means goal)
        t = (cx * dx + cy * dy) / norm_sq if norm_sq != 0 else 0
        # Candidate must be between the host and goal.
        if t < 0 or t > 1:
            continue

        # Calculate the projection point on the host-goal line.
        proj_x = host_position[0] + t * dx
        proj_y = host_position[1] + t * dy

        # Calculate the perpendicular deviation.
        deviation = math.hypot(candidate[0] - proj_x, candidate[1] - proj_y)
        if deviation > 10:
            continue

        # Condition 2: Check that the full circular area around the candidate is free.
        full_circle_free = True
        for key, occ in occupancy_dict.items():
            cell_x, cell_y  = key  # Only x and y are needed.
            if (cell_x - candidate[0]) ** 2 + (cell_y - candidate[1]) ** 2 <= radius ** 2:
                if occ != 0:
                    full_circle_free = False
                    break
        if full_circle_free:
            return candidate

    return None  # No valid candidate found.

def generate_traj_GF(env, episode_situation_holder, random_map_idx, geo_fence_spawn_step):
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    font_size = 15
    # Define color thresholds
    light_gray = 0.8  # Lightest gray (not white)
    dark_gray = 0.2  # Darkest gray (almost black)

    # plt.axvline(x=env.bound_collection[random_map_idx][0], c="green")
    # plt.axvline(x=env.bound_collection[random_map_idx][1], c="green")
    # plt.axhline(y=env.bound_collection[random_map_idx][2], c="green")
    # plt.axhline(y=env.bound_collection[random_map_idx][3], c="green")


    # draw occupied_poly
    for one_poly in env.world_map_2D_polyList_collection[random_map_idx][0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in env.world_map_2D_polyList_collection[random_map_idx][0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    # show geo-fence
    for geo_fence in env.geo_fence_area:
        fence_poly = shapelypoly_to_matpoly(geo_fence, False, 'red')  # the 3rd parameter is the edge color
        gf_x = geo_fence.centroid.x
        gf_y = geo_fence.centroid.y
        plt.text(gf_x, gf_y, f'{geo_fence_spawn_step} s')
        ax.add_patch(fence_poly)

    for agentIdx, agent in env.all_agents.items():
        dx = agent.goal[0][0] - agent.ini_pos[0]
        dy = agent.goal[0][1] - agent.ini_pos[1]
        angle = math.atan2(dy, dx)
        heading = math.degrees(angle)

        plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                 marker=MarkerStyle(">",
                                    fillstyle="right",
                                    transform=Affine2D().rotate_deg(heading)),
                 color='g', markersize=10, label='Origin')
        # plt.text(agent.ini_pos[0], agent.ini_pos[1], 'Origin')
        plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='r', markersize=10, label='Destination')
        # plt.text(agent.goal[-1][0], agent.goal[-1][1], 'Destination')


        # link individual drone's starting position with its goal
        ini = agent.ini_pos
        # for wp in agent.goal:
        ic = 0
        for wp in agent.ref_line.coords:
            # plt.plot(wp[0], wp[1], marker='*', color='y', markersize=10)
            if ic == 0:
                plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c', label='Reference Line')
            else:
                plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c')
            ic = ic + 1
            ini = wp
        trajectory = []
        which_to_see = -1
        for eps_time_step, ea_eps in enumerate(episode_situation_holder[which_to_see]):
            t = eps_time_step
            radius = 2.5
            gray_value = light_gray - t * (light_gray - dark_gray) / (len(episode_situation_holder[which_to_see]) - 1)
            color_value = grayscale_to_rgba(gray_value)
            x = ea_eps[0][0]
            y = ea_eps[0][1]
            trajectory.append([x, y])
            # ax.plot(x, y, 'o', color=plt.cm.gray(color_value), markersize=10)
            # circle = patches.Circle((x, y), 2.5, color=color_value, ec='black')
            circle = patches.Circle((x, y), radius, facecolor=color_value, edgecolor='white')
            ax.add_patch(circle)
            if eps_time_step == geo_fence_spawn_step:
                ax.text(x, y + 0.3, f't={eps_time_step}', fontsize=9, ha='center', va='center', color='black')

            # Plot large circle with dotted outline
            # outline_circle = patches.Circle((x, y), 7.5, facecolor='none', edgecolor=color_value, linestyle='dotted')
            # ax.add_patch(outline_circle)
        # Draw lines between the points
        trajectory = np.array(trajectory)
        ax.plot(trajectory[:, 0], trajectory[:, 1], color='black', linestyle='-', linewidth=1)

    # Adding titles and labels
    ax.set_title('Drone Trajectory with 3 geo-fences',fontsize=font_size)
    ax.set_xlabel('N-S direction (m)',fontsize=font_size)
    # ax.set_ylabel('E-W direction (m)',fontsize=font_size, labelpad=-8)  # use for random figure a
    ax.set_ylabel('E-W direction (m)',fontsize=font_size)
    # Customize tick size
    ax.tick_params(axis='both', which='major', labelsize=font_size)
    # Custom legend
    legend_elements = [
        Line2D([0], [0], linestyle='None', marker=MarkerStyle(">", fillstyle="right"), color='g', label='Origin', markerfacecolor='g', markersize=10),
        Line2D([0], [0], marker='*', color='w', label='Destination', markerfacecolor='r', markersize=15, linestyle='None'),
        Line2D([0], [0], color='cyan', lw=2, linestyle='dotted', label='Reference Line'),
        Line2D([0], [0], linestyle='None', marker='o', color='red', label='Geo-fence', markerfacecolor='none', markersize=10),
        Line2D([0], [0], linestyle='None', marker='o', color='white', label='Drone', markerfacecolor='black',
               markersize=10, alpha=0.5),
        Rectangle((0, 0), 1, 1, color='blue', alpha=0.9, label='Building Grids')  # Blue grid square for legend
    ]

    # Add legend
    # Variable to control the size of the square legend item
    square_size_factor = 0.9  # Adjust this value as needed

#     ax.legend(handles=legend_elements, loc='upper right', fontsize=font_size-2, borderpad=0.1, handler_map={
#     Rectangle: HandlerSquare(square_size_factor=square_size_factor)
# })

    plt.xlim(env.bound_collection[random_map_idx][0], env.bound_collection[random_map_idx][1]-5)
    plt.ylim(env.bound_collection[random_map_idx][2], env.bound_collection[random_map_idx][3])
    # plt.savefig('drone_trajectory.png')
    plt.show()


def generate_traj(env, episode_situation_holder, random_map_idx):
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    font_size = 15
    # Define color thresholds
    light_gray = 0.8  # Lightest gray (not white)
    dark_gray = 0.2  # Darkest gray (almost black)

    # plt.axvline(x=env.bound_collection[random_map_idx][0], c="green")
    # plt.axvline(x=env.bound_collection[random_map_idx][1], c="green")
    # plt.axhline(y=env.bound_collection[random_map_idx][2], c="green")
    # plt.axhline(y=env.bound_collection[random_map_idx][3], c="green")


    # draw occupied_poly
    for one_poly in env.world_map_2D_polyList_collection[random_map_idx][0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in env.world_map_2D_polyList_collection[random_map_idx][0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    # show geo-fence
    for geo_fence in env.geo_fence_area:
        fence_poly = shapelypoly_to_matpoly(geo_fence, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(fence_poly)

    for agentIdx, agent in env.all_agents.items():
        dx = agent.goal[0][0] - agent.ini_pos[0]
        dy = agent.goal[0][1] - agent.ini_pos[1]
        angle = math.atan2(dy, dx)
        heading = math.degrees(angle)

        plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                 marker=MarkerStyle(">",
                                    fillstyle="right",
                                    transform=Affine2D().rotate_deg(heading)),
                 color='g', markersize=10, label='Origin')
        # plt.text(agent.ini_pos[0], agent.ini_pos[1], 'Origin')
        plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='r', markersize=10, label='Destination')
        # plt.text(agent.goal[-1][0], agent.goal[-1][1], 'Destination')

        # link individual drone's starting position with its goal
        ini = agent.ini_pos
        # for wp in agent.goal:
        ic = 0
        for wp in agent.ref_line.coords:
            # plt.plot(wp[0], wp[1], marker='*', color='y', markersize=10)
            if ic == 0:
                plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c', label='Reference Line')
            else:
                plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c')
            ic = ic + 1
            ini = wp
        trajectory = []
        which_to_see = -1
        for eps_time_step, ea_eps in enumerate(episode_situation_holder[which_to_see]):
            t = eps_time_step
            radius = 2.5
            gray_value = light_gray - t * (light_gray - dark_gray) / (len(episode_situation_holder[which_to_see]) - 1)
            color_value = grayscale_to_rgba(gray_value)
            x = ea_eps[0][0]
            y = ea_eps[0][1]
            trajectory.append([x, y])
            # ax.plot(x, y, 'o', color=plt.cm.gray(color_value), markersize=10)
            # circle = patches.Circle((x, y), 2.5, color=color_value, ec='black')
            circle = patches.Circle((x, y), radius, facecolor=color_value, edgecolor='white')
            ax.add_patch(circle)
            # ax.text(x, y + 0.3, f't={eps_time_step}', fontsize=9, ha='center', va='center', color='black')

            # Plot large circle with dotted outline
            # outline_circle = patches.Circle((x, y), 7.5, facecolor='none', edgecolor=color_value, linestyle='dotted')
            # ax.add_patch(outline_circle)
        # Draw lines between the points
        trajectory = np.array(trajectory)
        ax.plot(trajectory[:, 0], trajectory[:, 1], color='black', linestyle='-', linewidth=1)

    # Adding titles and labels
    ax.set_title('Drone Trajectory with 3 geo-fences',fontsize=font_size)
    ax.set_xlabel('N-S direction (m)',fontsize=font_size)
    # ax.set_ylabel('E-W direction (m)',fontsize=font_size, labelpad=-8)  # use for random figure a
    ax.set_ylabel('E-W direction (m)',fontsize=font_size)
    # Customize tick size
    ax.tick_params(axis='both', which='major', labelsize=font_size)
    # Custom legend
    legend_elements = [
        Line2D([0], [0], linestyle='None', marker=MarkerStyle(">", fillstyle="right"), color='g', label='Origin', markerfacecolor='g', markersize=10),
        Line2D([0], [0], marker='*', color='w', label='Destination', markerfacecolor='r', markersize=15, linestyle='None'),
        Line2D([0], [0], color='cyan', lw=2, linestyle='dotted', label='Reference Line'),
        Line2D([0], [0], linestyle='None', marker='o', color='red', label='Geo-fence', markerfacecolor='none', markersize=10),
        Line2D([0], [0], linestyle='None', marker='o', color='white', label='Drone', markerfacecolor='black',
               markersize=10, alpha=0.5),
        Rectangle((0, 0), 1, 1, color='blue', alpha=0.9, label='Building Grids')  # Blue grid square for legend
    ]

    # Add legend
    # Variable to control the size of the square legend item
    square_size_factor = 0.9  # Adjust this value as needed

#     ax.legend(handles=legend_elements, loc='upper right', fontsize=font_size-2, borderpad=0.1, handler_map={
#     Rectangle: HandlerSquare(square_size_factor=square_size_factor)
# })

    plt.xlim(env.bound_collection[random_map_idx][0], env.bound_collection[random_map_idx][1]-5)
    plt.ylim(env.bound_collection[random_map_idx][2], env.bound_collection[random_map_idx][3])
    # plt.savefig('drone_trajectory.png')
    plt.show()


def initialize_excel_file(file_path):
    # Create a new workbook and add three empty sheets
    wb = Workbook()
    wb.save(file_path)


def append_to_excel(file_path, data):
    # Try to load the workbook, and create it if it does not exist
    try:
        wb = load_workbook(file_path)
        new_workbook = False
    except FileNotFoundError:
        wb = Workbook()
        new_workbook = True
        # wb.remove(wb.active)  # Remove the default sheet created by openpyxl if not need
    except Exception as e:  # Catch other exceptions, such as invalid file format
        print(f"An error occurred: {e}")
        return
    if new_workbook:
        if 'Sheet' in wb.sheetnames:
            # If the sheet exists, remove it
            wb.remove('Sheet')
    # Check if the required sheets exist, if not create them
    if len(data) == 3:  # for record eps time
        number_of_sheets = len(data)
        required_sheets = ['Sheet'+str(i) for i in range(number_of_sheets)]
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(title=sheet_name)

        # Append the data to the respective sheets
        for i, portion in enumerate(data):
            sheet = wb[required_sheets[i]]
            if isinstance(portion, list):
                for item in portion:
                    sheet.append(item if isinstance(item, list) else [item])
                # If this is the last portion of data, append -1 to the last row
                if i == len(data) - 1:
                    sheet.append([-1])
            else:
                sheet.append([portion])
    else:
        if isinstance(data[0], list):  # for record reward
            # In reward record, we are recording in the form that one step have 3 agents
            required_sheets = ["agent0", "agent1", "agent2"]
            for sheet_name in required_sheets:
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(title=sheet_name)
            for idx, step_reward in enumerate(data):
                for agent_idx, single_sheet_name in enumerate(required_sheets):
                    sheet = wb[single_sheet_name]
                    sheet.append(step_reward[agent_idx])
                    if idx == len(data)-1:
                        sheet.append([-9999])

        else:  # for record noise
            wb.create_sheet(title='noise given at each step')
            sheet = wb['noise given at each step']
            for i, portion in enumerate(data):
                sheet.append(list(portion))
            sheet.append([-999])

    # Save the workbook
    wb.save(file_path)


def animate(frame_num, ax, env, trajectory_eachPlay, random_map_idx):
    ax.clear()
    plt.axis('equal')
    plt.xlim(env.bound_collection[random_map_idx][0], env.bound_collection[random_map_idx][1])
    plt.ylim(env.bound_collection[random_map_idx][2], env.bound_collection[random_map_idx][3])
    plt.axvline(x=env.bound_collection[random_map_idx][0], c="green")
    plt.axvline(x=env.bound_collection[random_map_idx][1], c="green")
    plt.axhline(y=env.bound_collection[random_map_idx][2], c="green")
    plt.axhline(y=env.bound_collection[random_map_idx][3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")

    # draw occupied_poly
    for one_poly in env.world_map_2D_polyList_collection[random_map_idx][0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in env.world_map_2D_polyList_collection[random_map_idx][0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    # show geo-fence
    for geo_fence in env.geo_fence_area:
        fence_poly = shapelypoly_to_matpoly(geo_fence, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(fence_poly)

    for agentIdx, agent in env.all_agents.items():
        plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                 marker=MarkerStyle(">",
                                    fillstyle="right",
                                    transform=Affine2D().rotate_deg(math.degrees(agent.heading))),
                 color='y')
        plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
        plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='y', markersize=10)
        plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)

        # link individual drone's starting position with its goal
        ini = agent.ini_pos
        # for wp in agent.goal:
        for wp in agent.ref_line.coords:
            # plt.plot(wp[0], wp[1], marker='*', color='y', markersize=10)
            plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c')
            ini = wp

    for a_idx, agent in enumerate(trajectory_eachPlay[frame_num]):
        x, y = agent[0], agent[1]
        # plot drone's own protective circle without filling
        # plt.plot(x, y, 'o', color='r')

        # plt.text(x-1, y-1, 'agent_'+str(a_idx)+'_'+str(round(float(frame_num), 2)))
        plt.text(x-1, y-1, 'agent_'+str(a_idx)+'_'+str(agent[2]))

        # plot the heading of the current drone
        heading_radians = agent[3]['A0_heading']
        arrow_length = 2.5
        arrow_end_x = x + arrow_length * np.cos(heading_radians)
        arrow_end_y = y + arrow_length * np.sin(heading_radians)
        ax.annotate('', xy=(arrow_end_x, arrow_end_y), xytext=(x, y),
                    arrowprops=dict(arrowstyle='->', lw=2, color='blue'))
        # plot the nearest point on line from the drone's position
        nearest_pt = agent[3]['neareset_point']
        nearest_dist = agent[3]['deviation_to_ref_line']
        nearest_dist_rw = agent[3]['deviation_to_ref_line_reward']
        plt.scatter(nearest_pt.x, nearest_pt.y, color='g')
        plt.text(nearest_pt.x, nearest_pt.y, str(nearest_dist) + '_' + str(nearest_dist_rw))

        # plot the drone's detection range
        detec_circle = Point(x, y).buffer(30 / 2, cap_style='round')
        detec_circle_mat = shapelypoly_to_matpoly(detec_circle, inFill=False, Edgecolor='g')
        ax.add_patch(detec_circle_mat)

        # plot the shortest radar prob
        shortest_distance_element = min(agent[3]['A0_observable space'], key=lambda x: x[0])
        if shortest_distance_element[0] < 15:
            plt.plot([x, shortest_distance_element[2].x], [y, shortest_distance_element[2].y], linestyle='dashed',
                     color='b')

        self_circle = Point(x, y).buffer(env.all_agents[0].protectiveBound, cap_style='round')
        grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')
        ax.add_patch(grid_mat_Scir)

    return ax.patches + [ax.texts]


def get_history_tensor(history, sequence_length, input_size):
    # If history is shorter than the sequence length, pad it with zeros
    if len(history) < sequence_length:
        padding = torch.zeros(sequence_length - len(history), input_size)
        history_tensor = torch.cat((padding, torch.stack(list(history))))
    else:
        history_tensor = torch.stack(list(history))

    # Add a batch dimension and return
    return history_tensor.unsqueeze(0)


def save_gif(env, trajectory_eachPlay, pre_fix, episode_to_check, episode, random_map_idx):
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)

    plt.axis('equal')
    plt.xlim(env.bound_collection[random_map_idx][0], env.bound_collection[random_map_idx][1])
    plt.ylim(env.bound_collection[random_map_idx][2], env.bound_collection[random_map_idx][3])
    plt.axvline(x=env.bound_collection[random_map_idx][0], c="green")
    plt.axvline(x=env.bound_collection[random_map_idx][1], c="green")
    plt.axhline(y=env.bound_collection[random_map_idx][2], c="green")
    plt.axhline(y=env.bound_collection[random_map_idx][3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")

    # draw occupied_poly
    for one_poly in env.world_map_2D_polyList_collection[random_map_idx][0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in env.world_map_2D_polyList_collection[random_map_idx][0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    for agentIdx, agent in env.all_agents.items():
        plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                 marker=MarkerStyle(">",
                                    fillstyle="right",
                                    transform=Affine2D().rotate_deg(math.degrees(agent.heading))),
                 color='y')
        plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
        # plot self_circle of the drone
        self_circle = Point(agent.ini_pos[0],
                            agent.ini_pos[1]).buffer(agent.protectiveBound, cap_style='round')
        grid_mat_Scir = shapelypoly_to_matpoly(self_circle, inFill=False, Edgecolor='k')
        ax.add_patch(grid_mat_Scir)

        # plot drone's detection range
        detec_circle = Point(agent.ini_pos[0],
                             agent.ini_pos[1]).buffer(agent.detectionRange / 2, cap_style='round')
        detec_circle_mat = shapelypoly_to_matpoly(detec_circle, inFill=False, Edgecolor='g')
        ax.add_patch(detec_circle_mat)

        plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='y', markersize=10)
        plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)

    # Create animation
    ani = animation.FuncAnimation(fig, animate, fargs=(ax, env, trajectory_eachPlay, random_map_idx), frames=len(trajectory_eachPlay),
                                  interval=300, blit=False)
    # Save as GIF
    gif_path = pre_fix + '\episode_' + episode_to_check + 'simulation_num_' + str(episode) + '.gif'
    ani.save(gif_path, writer='pillow')

    # Close figure
    plt.close(fig)


def view_static_traj(env, trajectory_eachPlay, random_map_idx):
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    # display initial condition
    # global_state = env.reset_world(show=0)  # just a dummy to reset all condition so that initial condition can be added to the output graph
    for agentIdx, agent in env.all_agents.items():
        # if agentIdx != 0:
        #     continue
        plt.plot(agent.ini_pos[0], agent.ini_pos[1],
                 marker=MarkerStyle(">",
                                    fillstyle="right",
                                    transform=Affine2D().rotate_deg(math.degrees(agent.heading))),
                 color='y')
        plt.text(agent.ini_pos[0], agent.ini_pos[1], agent.agent_name)
        # plot self_circle of the drone
        self_circle = Point(agent.ini_pos[0],
                            agent.ini_pos[1]).buffer(agent.protectiveBound, cap_style='round')
        grid_mat_Scir = shapelypoly_to_matpoly(self_circle, inFill=False, Edgecolor='k')
        ax.add_patch(grid_mat_Scir)

        # plot drone's detection range
        detec_circle = Point(agent.ini_pos[0],
                             agent.ini_pos[1]).buffer(agent.detectionRange / 2, cap_style='round')
        detec_circle_mat = shapelypoly_to_matpoly(detec_circle, inFill=False, Edgecolor='g')
        ax.add_patch(detec_circle_mat)

        # link individual drone's starting position with its goal
        ini = agent.ini_pos
        # for wp in agent.goal:
        for wp in agent.ref_line.coords:
            plt.plot(wp[0], wp[1], marker='*', color='y', markersize=10)
            plt.plot([wp[0], ini[0]], [wp[1], ini[1]], '--', color='c')
            # plot drone's detection range
            wp_circle = Point(wp[0], wp[1]).buffer(agent.protectiveBound, cap_style='round')
            wp_circle_mat = shapelypoly_to_matpoly(wp_circle, inFill=False, Edgecolor='g')
            ax.add_patch(wp_circle_mat)
            ini = wp

        plt.plot(agent.goal[-1][0], agent.goal[-1][1], marker='*', color='y', markersize=10)
        plt.text(agent.goal[-1][0], agent.goal[-1][1], agent.agent_name)

    # draw trajectory in current episode
    for trajectory_idx, trajectory_val in enumerate(trajectory_eachPlay):  # each time step
        # if trajectory_idx >= 10:
        #     break
        for agentIDX, each_agent_traj in enumerate(trajectory_val):  # for each agent's motion in a time step
            # if agentIDX != 0:
            #     continue
            x, y = each_agent_traj[0], each_agent_traj[1]
            plt.plot(x, y, 'o', color='r')
            # detec_circle = Point(x, y).buffer(15, cap_style='round')
            # detec_circle_mat = shapelypoly_to_matpoly(detec_circle, inFill=False, Edgecolor='purple')
            # ax.add_patch(detec_circle_mat)
            # plot the heading of the current drone
            heading_radians = each_agent_traj[3]['A0_heading']
            arrow_length = 2.5
            arrow_end_x = x + arrow_length * np.cos(heading_radians)
            arrow_end_y = y + arrow_length * np.sin(heading_radians)
            ax.annotate('', xy=(arrow_end_x, arrow_end_y), xytext=(x, y),
                        arrowprops=dict(arrowstyle='->', lw=2, color='blue'))
            # plot the nearest point on line from the drone's position
            nearest_pt = each_agent_traj[3]['neareset_point']
            nearest_dist = each_agent_traj[3]['deviation_to_ref_line']
            nearest_dist_rw = each_agent_traj[3]['deviation_to_ref_line_reward']
            # plt.scatter(nearest_pt.x, nearest_pt.y, color='g')
            # plt.text(nearest_pt.x, nearest_pt.y, str(round(nearest_dist, 3))+'_'+str(round(nearest_dist_rw, 3)))

            # plot the shortest radar prob
            shortest_distance_element = min(each_agent_traj[3]['A0_observable space'], key=lambda x: x[0])
            if shortest_distance_element[0] < 15:
                plt.plot([x, shortest_distance_element[2].x], [y, shortest_distance_element[2].y], linestyle='dashed', color='b')

            # plt.text(x-1, y-1, str(round(float(reward_each_agent[trajectory_idx][agentIDX]),2)))
            # plt.text(475, 400-trajectory_idx, 'agent' + str(agentIDX) + '_' + str(each_agent_traj[2].round(3)))
            plt.text(env.bound_collection[random_map_idx][0], env.bound_collection[random_map_idx][3]-trajectory_idx, 'step' + str(trajectory_idx) + '_GL_' + str( each_agent_traj[3]['goal_leading_reward'].round(3)))
            # plt.text(x - 1, y - 1, 'agent_' + str(agentIDX) + '_' + str(each_agent_traj[2]))
            self_circle = Point(x, y).buffer(env.all_agents[0].protectiveBound, cap_style='round')
            grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')
            ax.add_patch(grid_mat_Scir)

    # show geo-fence
    for geo_fence in env.geo_fence_area:
        fence_poly = shapelypoly_to_matpoly(geo_fence, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(fence_poly)

    # draw occupied_poly
    for one_poly in env.world_map_2D_polyList_collection[random_map_idx][0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in env.world_map_2D_polyList_collection[random_map_idx][0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    plt.axis('equal')
    plt.xlim(env.bound_collection[random_map_idx][0], env.bound_collection[random_map_idx][1])
    plt.ylim(env.bound_collection[random_map_idx][2], env.bound_collection[random_map_idx][3])
    plt.axvline(x=env.bound_collection[random_map_idx][0], c="green")
    plt.axvline(x=env.bound_collection[random_map_idx][1], c="green")
    plt.axhline(y=env.bound_collection[random_map_idx][2], c="green")
    plt.axhline(y=env.bound_collection[random_map_idx][3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")
    plt.show()


def total_length_to_end_of_line_without_cross(initial_point, linestring):
    """
    Calculate the total distance from an initial point to the nearest point on the line and
    from there to the end of the line.

    Parameters:
    initial_point (tuple): The initial point as (x, y).
    linestring (LineString): The LineString object.

    Returns:
    float: The total distance from the initial point to the end of the LineString.
    """
    # Create a Point object from the tuple
    point = Point(initial_point)

    # Find the nearest point on the line to the initial point
    nearest_point_on_line = linestring.interpolate(linestring.project(point))

    # Calculate the distance from the nearest point on the line to the end of the line
    projected_distance = linestring.project(nearest_point_on_line)
    distance_to_end_of_line = linestring.length - projected_distance

    return distance_to_end_of_line


def compute_projected_velocity(vehicle_velocity, reference_path, vehicle_position):
    # Find the closest point on the path to the vehicle's current position
    closest_point = nearest_points(reference_path, vehicle_position)[0]
    # Check if the closest point is a turning point
    turning_points = [Point(p) for p in reference_path.coords[1:-1]]  # Exclude the first and last points
    # If closest_point is a turning point, use the outgoing path segment for tangent calculation
    if closest_point in turning_points:
        # Find the index of the turning point in the list of points
        turning_index = turning_points.index(closest_point)

        # The outgoing segment starts from the turning point
        segment_start = np.array(reference_path.coords[turning_index + 1])
        segment_end = np.array(reference_path.coords[turning_index + 2])

        # Calculate the tangent vector (unit vector) of the segment
        tangent_vector = (segment_end - segment_start) / np.linalg.norm(segment_end - segment_start)

        # # Project the vehicle's velocity onto the tangent vector
        # projected_velocity = np.dot(vehicle_velocity, tangent_vector) * tangent_vector

        # output dot value only.
        projected_velocity = np.dot(vehicle_velocity, tangent_vector)  # this is a scalar value.
    else:
        # If it's not a turning point, just find the index of the closest segment
        distance_along_line = reference_path.project(closest_point)
        # Find the segment that contains this distance
        cumulative_distance = 0
        for i in range(len(reference_path.coords) - 1):
            segment_start = np.array(reference_path.coords[i])
            segment_end = np.array(reference_path.coords[i + 1])
            segment_length = LineString([segment_start, segment_end]).length
            cumulative_distance += segment_length
            if cumulative_distance > distance_along_line:
                break
        # Compute the tangent vector as before
        tangent_vector = (segment_end - segment_start) / np.linalg.norm(segment_end - segment_start)
        # Project the vehicle's velocity onto the tangent vector
        # projected_velocity = np.dot(vehicle_velocity, tangent_vector) * tangent_vector

        # output dot value only.
        projected_velocity = np.dot(vehicle_velocity, tangent_vector)  # this is a scalar value.

    return projected_velocity


def total_length_to_end_of_line(initial_point, linestring):
   """
   Calculate the total distance from an initial point to the nearest point on the line and
   from there to the end of the line.
   Parameters:
   initial_point (tuple): The initial point as (x, y).
   linestring (LineString): The LineString object.
   Returns:
   float: The total distance from the initial point to the end of the LineString.
   """
   # Create a Point object from the tuple
   point = Point(initial_point)
   # Find the nearest point on the line to the initial point
   nearest_point_on_line = linestring.interpolate(linestring.project(point))
   # Calculate the distance from the initial point to the nearest point on the line
   distance_to_line = point.distance(nearest_point_on_line)
   # Calculate the distance from the nearest point on the line to the end of the line
   projected_distance = linestring.project(nearest_point_on_line)
   distance_to_end_of_line = linestring.length - projected_distance
   # Sum the distances to get the total distance
   total_distance = distance_to_line + distance_to_end_of_line
   return total_distance


def sort_polygons(polygons):  # this sorting is left to right, but bottom to top. so, 0th is below 2nd. [[2,3],[0,1]]
    boxes = [polygon.bounds for polygon in polygons]
    sorted_boxes = sorted(boxes, key=lambda box: (box[1], box[0]))
    sorted_polygons = [polygons[boxes.index(box)] for box in sorted_boxes]
    return sorted_polygons


# this is to remove the need for the package descrete
def shapelypoly_to_matpoly(ShapelyPolgon, inFill=False, Edgecolor='black', FcColor='blue'):
    xcoo, ycoo = ShapelyPolgon.exterior.coords.xy
    matPolyConverted = matPolygon(xy=list(zip(xcoo, ycoo)), fill=inFill, edgecolor=Edgecolor, facecolor=FcColor)
    return matPolyConverted


def extract_individual_obs(combine_state, agent_idx):
    individual_obs = []
    self_obs = combine_state[agent_idx][0]
    self_obs_grid = combine_state[agent_idx][1]
    self_surround = combine_state[agent_idx][2]
    individual_obs = [self_obs, self_obs_grid, self_surround]
    # return individual_obs  # for old V5
    return combine_state[agent_idx]


def map_range(value, coe_a):
    # Calculate the normalized value
    out_min = -coe_a
    out_max = coe_a
    # tanh() has range of -1 to 1
    in_min = -1
    in_max = 1
    normalized = (value - in_min) / (in_max - in_min)
    # Map the normalized value to the output range
    mapped = out_min + (normalized * (out_max - out_min))
    # Return the mapped value
    return mapped


def compute_potential_conflict(cur_drone_pos, cur_drone_vel, cur_drone_protRad, cur_neigh_pos, cur_neigh_vel,
                               cur_neigh_protRad, cur_neigh_idx, current_ts):
    minus_rel_dist_before = -1 * (cur_drone_pos - cur_neigh_pos)  # always current drone - neighbours
    rel_vel_before = (cur_drone_vel - cur_neigh_vel)
    if not (np.any(cur_drone_vel) or np.any(cur_neigh_vel)):
        # print("All elements in both arrays are zero")
        rel_vel_before = rel_vel_before + 1e-10  # add a very small number to suppress division by zero warning

    rel_vel_SQnorm_before = np.square(np.linalg.norm(rel_vel_before))

    t_cpa_before = np.dot(minus_rel_dist_before, rel_vel_before) / rel_vel_SQnorm_before
    d_cpa_before = np.linalg.norm(((cur_drone_pos - cur_neigh_pos) + (rel_vel_before * t_cpa_before)))
    # time to potential conflict, t_cpa is the actual time in sec, so, if t_cap=3, meaning in order to reach d_cap, the decision making agent need to take 6 steps. Currently our
    # if (t_cpa_before >= 0) and (t_cpa_before <= 1.5) and (d_cpa_before < (cur_drone_protRad + cur_neigh_protRad)):
    #     # pc_list.append(cur_neigh_idx)
    #     pc_list[cur_neigh_idx] = [t_cpa_before, d_cpa_before]
        # double check
        # host_future_pos = cur_drone_pos + (t_cpa_before*cur_drone_vel)
        # intru_future_pos = cur_neigh_pos + (t_cpa_before*cur_neigh_vel)
        # doubleCheck_dcpa = np.linalg.norm((host_future_pos-intru_future_pos))
    if (t_cpa_before >= 0) and (t_cpa_before < 3):
        return [cur_neigh_idx, t_cpa_before, d_cpa_before]
    else:
        return []


def padding_list(max_grid_obs_dim, input_list):
    # padding actions
    if isinstance(input_list, np.float64):
        print("check")
    tobePad_gridObs = list(np.zeros(max_grid_obs_dim - len(input_list), dtype=int))
    padded_gridObs = input_list + tobePad_gridObs
    return padded_gridObs


def preprocess_batch_for_critic_net(input_state, batch_size):
    critic_own_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents
    critic_grid_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents
    critic_neigh_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents
    for batch_idx in range(batch_size):
        critic_own_cur_state = []
        critic_own_grid_state = []
        critic_own_neigh_state = []
        for agent_cur in input_state:
            critic_own_cur_state.append(agent_cur[0][batch_idx, :])
            critic_own_grid_state.append(agent_cur[1][batch_idx, :])
            # for neigh, first: max_nei_num X single_nei_features flatten to 1D array
            flat_nei = agent_cur[2][batch_idx, :].flatten()  # default is flatten in a row.
            critic_own_neigh_state.append(flat_nei)
        critic_own_batched_cur_state.append(np.array(critic_own_cur_state).reshape((1, -1)))
        critic_grid_batched_cur_state.append(np.array(critic_own_grid_state).reshape((1, -1)))
        critic_neigh_batched_cur_state.append(np.array(critic_own_neigh_state).reshape((1, -1)))

    cur_state_pre_processed = [T.tensor(np.array(critic_own_batched_cur_state)),
                               T.tensor(np.array(critic_grid_batched_cur_state)),
                               T.tensor(np.array(critic_neigh_batched_cur_state))]
    return cur_state_pre_processed


def preprocess_batch_for_critic_net_v2(input_state, batch_size):
    critic_own_batched_cur_state = []  # batch_size X one_agent_feature * max_num_agents

    for batch_idx in range(batch_size):
        critic_own_cur_state = []

        for agent_cur in input_state:
            critic_own_cur_state.append(agent_cur[batch_idx, :])




        critic_own_batched_cur_state.append(np.array(critic_own_cur_state).reshape((1, -1)))


    cur_state_pre_processed = T.tensor(np.array(critic_own_batched_cur_state))  # batch X (1 x no_agent x feature size)

    return cur_state_pre_processed


# class OUNoise:
#
#     def __init__(self, action_dimension, largest_Nsigma, smallest_Nsigma, ini_sigma, mu=0, theta=0.15):  # sigma is the initial magnitude of the OU_noise
#         self.action_dimension = action_dimension
#         self.mu = mu
#         self.theta = theta
#         self.sigma = ini_sigma
#         self.largest_sigma = largest_Nsigma
#         self.smallest_sigma = smallest_Nsigma
#         self.state = np.ones(self.action_dimension) * self.mu
#         self.reset()
#
#     def reset(self):
#         self.state = np.ones(self.action_dimension) * self.mu
#
#     def noise(self):
#         x = self.state
#         dx = self.theta * (self.mu - x) + self.sigma * np.random.randn(len(x))
#         self.state = x + dx
#         return self.state

class OUActionNoise:
    def __init__(self, mu, sigma, theta, dt):
        self.ou_gen = None
        self.sigma = sigma
        self.mu = mu
        self.theta = theta
        self.dt = dt

    def reset(self):
        self.ou_gen = ornstein_uhlenbeck_gen(self.mu, self.sigma, self.theta, self.dt)

    def __call__(self):
        return next(self.ou_gen)


def action_rescaler(low, high):
    halfspan = (high - low) / 2.0
    middle = (high + low) / 2.0

    def rescaler(action):
        return np.clip((action + middle) * halfspan, low, high)

    return rescaler


def ornstein_uhlenbeck_gen(
    mu: np.ndarray = None,
    sigma: float = 1.0,
    theta: float = 0.15,
    dt: float = 1e-2,
):
    rescaler = action_rescaler(-1, 1)  # this acts like an noise clipping, align with the paper's description
    noise_gen = ornstein_uhlenbeck_unscaled(mu=mu, sigma=sigma, theta=theta, dt=dt)
    for action in noise_gen:
        yield rescaler(action)

def ornstein_uhlenbeck_unscaled(
    mu: np.ndarray = None,
    sigma: float = 1.0,
    dim: int = None,
    theta: float = 0.15,
    dt: float = 1e-2,
):
    if mu is None:
        if not dim:
            raise ValueError("Either mu or dimensionality parameter dim must be set!")
        mu = np.zeros(dim)
    last_noise = np.zeros_like(mu)
    while True:
        noise = (
            last_noise
            + theta * (mu - last_noise) * dt
            + sigma * np.sqrt(dt) * np.random.normal(size=mu.shape)
        )
        last_noise = noise
        yield noise

class NormalizeData:
    def __init__(self, x_min_max, y_min_max, spd_max, acc_range):
        self.normalize_max = 1
        self.normalize_min = -1
        self.dis_min_x = x_min_max[0]
        self.dis_max_x = x_min_max[1]
        self.dis_min_y = y_min_max[0]
        self.dis_max_y = y_min_max[1]
        self.spd_max = spd_max
        self.acc_min = acc_range[0]
        self.acc_max = acc_range[1]
        self.scale_attribute()

    def scale_attribute(self):
        self.x_scale = (self.normalize_max-self.normalize_min)/(self.dis_max_x - self.dis_min_x)
        self.y_scale = (self.normalize_max-self.normalize_min)/(self.dis_max_y - self.dis_min_y)

    def nmlz_pos(self, pos_c):
        x, y = pos_c[0], pos_c[1]
        x_normalized = 2 * ((x - self.dis_min_x) / (self.dis_max_x - self.dis_min_x)) - 1
        y_normalized = 2 * ((y - self.dis_min_y) / (self.dis_max_y - self.dis_min_y)) - 1
        return np.array([x_normalized, y_normalized])

    def scale_pos(self, pos_c):  # NOTE: this method is essentially same as min-max normalize approach, but we need this appraoch to calculate x & y scale
        x_normalized = self.normalize_min + (pos_c[0] - self.dis_min_x) * self.x_scale
        y_normalized = self.normalize_min + (pos_c[1] - self.dis_min_y) * self.y_scale
        return np.array([x_normalized, y_normalized])

    def scale_vel(self, change_in_pos):
        return np.array([self.x_scale * change_in_pos[0], self.y_scale * change_in_pos[1]])

    def nmlz_pos_diff(self, diff):
        dx, dy = diff[0], diff[1]
        dx_min = self.dis_min_x-self.dis_max_x
        dx_max = self.dis_max_x-self.dis_min_x
        dy_min = self.dis_min_y-self.dis_max_y
        dy_max = self.dis_max_y-self.dis_min_y
        dx_normalized = 2 * ((dx - dx_min) / (dx_max - dx_min)) - 1
        dy_normalized = 2 * ((dy - dy_min) / (dy_max - dy_min)) - 1
        return dx_normalized, dy_normalized

    def nmlz_vel(self, cur_vel):
        vx, vy = cur_vel[0], cur_vel[1]
        vx_normalized = vx / self.spd_max
        vy_normalized = vy / self.spd_max
        # vx_normalized = (vx / self.spd_max) * 2 - 1
        # vy_normalized = (vy / self.spd_max) * 2 - 1
        return np.array([vx_normalized, vy_normalized])

    def nmlz_acc(self, cur_acc):
        ax, ay = cur_acc[0], cur_acc[1]
        ax_normalized = ((ax - self.acc_min) / (self.acc_max-self.acc_min)) * 2 - 1
        ay_normalized = (ay - self.acc_min) / (self.acc_max-self.acc_min) * 2 - 1
        return ax_normalized, ay_normalized


def BetaNoise(action, noise_scale):
    action = action.detach().numpy()  # since the input is a tensor we must convert it to numpy before operations
    sign = np.sign(action)  # tracking the sign so we can flip the samples later
    action = abs(action)  # we only use right tail of beta
    alpha = 1 / noise_scale  # this determines the how contentrated the beta dsn is
    value = 0.5 + action / 2  # converting from action space of -1 to 1 to beta space of 0 to 1
    beta = alpha * (1 - value) / value  # calculating beta
    beta = beta + 1.0 * (
                (alpha - beta) / alpha)  # adding a little bit to beta prevents actions getting stuck at -1 or 1
    sample = np.random.beta(alpha, beta)  # sampling from the beta distribution
    sample = sign * sample + (1 - sign) / 2  # flipping sample if sign is <0 since we only use right tail of beta dsn

    action_output = 2 * sample - 1  # converting back to action space -1 to 1
    return torch.tensor(action_output)  # converting back to tensor


def GaussNoise(action, noise_scale):
    n = np.random.normal(0, 1, len(action))  # create some standard normal noise
    return torch.clamp(action + torch.tensor(noise_scale * n).float(), -1, 1)  # add the noise to the actions


def WeightedNoise(action, noise_scale, action_type):
    """
    Returns the epsilon scaled noise distribution for adding to Actor
    calculated action policy.
    """
    if action_type == 'continuous':
        target = np.random.uniform(-1, 1, 2)  # the action space is -1 to 1
    elif action_type == 'discrete':
        target = np.random.uniform(0, 1, 4)  # action space is discrete
        target = target / sum(target)
    action = noise_scale * target + (
                1 - noise_scale) * action.detach().numpy()  # take a weighted average with noise_scale as the noise weight
    return torch.tensor(action).float()

def display_trajectory(cur_env, combined_trajectory, eps_to_watch):
    episode_to_show = eps_to_watch
    episode_steps = combined_trajectory[episode_to_show]
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    # # draw link towards destination for all drones, destination for each drone didn't change
    for agentIdx in cur_env.all_agents:
        plt.plot([cur_env.all_agents[agentIdx].ini_pos[0], cur_env.all_agents[agentIdx].goal[0][0]],
                 [cur_env.all_agents[agentIdx].ini_pos[1], cur_env.all_agents[agentIdx].goal[0][1]], '--', color='c')
    reward, done = [], []
    crash_penalty = -100
    reach_target = 1000
    potential_conflict_count = 0
    fixed_domino_reward = 1
    x_left_bound = LineString([(cur_env.bound[0], -9999), (cur_env.bound[0], 9999)])
    x_right_bound = LineString([(cur_env.bound[1], -9999), (cur_env.bound[1], 9999)])
    y_bottom_bound = LineString([(-9999, cur_env.bound[2]), (9999, cur_env.bound[2])])
    y_top_bound = LineString([(-9999, cur_env.bound[3]), (9999, cur_env.bound[3])])

    allBuildingSTR = STRtree(cur_env.world_map_2D_polyList[0][0])

    for step_idx, agents_traj in enumerate(episode_steps):
        step_R = []
        step_D = []
        for ea_idx, each_agent in enumerate(agents_traj):
            # propagate environment
            cur_env.all_agents[ea_idx].pre_pos = cur_env.all_agents[ea_idx].pos
            cur_env.all_agents[ea_idx].pos = np.array([each_agent[0], each_agent[1]])
            # check reward for each step
            # calculate the deviation from the reference path after an action has been taken
            curPoint = Point(cur_env.all_agents[ea_idx].pos)
            host_refline = LineString([cur_env.all_agents[ea_idx].ini_pos, cur_env.all_agents[ea_idx].goal[0]])
            cross_track_deviation = curPoint.distance(host_refline)  # deviation from the reference line, cross track error

            host_pass_line = LineString([cur_env.all_agents[ea_idx].pre_pos, cur_env.all_agents[ea_idx].pos])
            host_passed_volume = host_pass_line.buffer(cur_env.all_agents[ea_idx].protectiveBound, cap_style='round')
            possiblePoly = allBuildingSTR.query(host_passed_volume)
            for element in possiblePoly:
                if allBuildingSTR.geometries.take(element).intersection(host_passed_volume):
                    collide_building = 1
                    print("drone_{} crash into building when moving from {} to {} at time step {}".format(ea_idx,
                                                                                                          cur_env.all_agents[
                                                                                                              ea_idx].pre_pos,
                                                                                                          cur_env.all_agents[
                                                                                                              ea_idx].pos,
                                                                                                          step_idx))
                    break

            tar_circle = Point(cur_env.all_agents[ea_idx].goal[0]).buffer(1, cap_style='round')
            goal_cur_intru_intersect = host_passed_volume.intersection(tar_circle)
            if not goal_cur_intru_intersect.is_empty:  # reached goal?
                print("drone_{} has reached its goal at time step {}".format(ea_idx, step_idx))
                step_D.append(True)
                step_R.append(np.array(reach_target))
            # exceed bound condition, don't use current point, use current circle or else will have condition that
            elif x_left_bound.intersects(host_passed_volume) or x_right_bound.intersects(host_passed_volume) or y_bottom_bound.intersects(host_passed_volume) or y_top_bound.intersects(host_passed_volume):
                print("drone_{} has crash into boundary at time step {}".format(ea_idx, step_idx))
                step_R.append(np.array(crash_penalty))
                step_D.append(True)
            # exceed bound or crash into buildings or crash with other neighbors
            elif collide_building == 1 or len(collision_drones) > 0:
                reward.append(np.array(crash_penalty))
                done.append(True)
            else:  # a normal step taken
                step_D.append(False)
                crossCoefficient = 1
                goalCoefficient = 6
                # cross track error term
                cross_track_error = (20 / ((cross_track_deviation * cross_track_deviation) / 200 + 1)) - 3.5
                # Distance between drone and its goal for two consecutive time step
                before_dist_hg = np.linalg.norm(cur_env.all_agents[ea_idx].pre_pos - cur_env.all_agents[ea_idx].goal[0])
                after_dist_hg = np.linalg.norm(cur_env.all_agents[ea_idx].pos - cur_env.all_agents[ea_idx].goal[0])  # distance to goal after action
                delta_hg = goalCoefficient * (before_dist_hg - after_dist_hg)
                # a small penalty for discourage the agent to stay in one single spot
                if (before_dist_hg - after_dist_hg) <= 2:
                    small_step_penalty = 50
                else:
                    small_step_penalty = 0
                # Domino term also use as an indicator for agent to avoid other drones. so no need to specifically
                # add a term to avoid surrounding drones
                # step_reward = crossCoefficient*cross_track_error + delta_hg + dominoTerm - small_step_penalty
                step_reward = crossCoefficient*cross_track_error + delta_hg - small_step_penalty
                # step_reward = delta_hg
                # convert to arr
                step_R.append(np.array(step_reward))
                plt.text(each_agent[0] + 5, each_agent[1], str(np.array(round(step_reward,1))))

            # plot self_circle of the drone
            self_circle = Point(each_agent[0], each_agent[1]).buffer(2.5, cap_style='round')
            grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')

            # label drone time step for the position
            plt.text(each_agent[0], each_agent[1], str(ea_idx))
            #plt.text(each_agent[0]+5, each_agent[1], str(step_idx))

            ax.add_patch(grid_mat_Scir)
        reward.append(step_R)
        done.append(step_D)
    # print reward
    print_R = 0
    for eps_stepR in reward:
        print_R = print_R + sum(eps_stepR)
    print(print_R)

    # draw occupied_poly
    for one_poly in cur_env.world_map_2D_polyList[0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'y', 'b')
        ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in cur_env.world_map_2D_polyList[0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'y')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in cur_env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    plt.axis('equal')
    plt.xlim(cur_env.bound[0], cur_env.bound[1])
    plt.ylim(cur_env.bound[2], cur_env.bound[3])
    plt.axvline(x=cur_env.bound[0], c="green")
    plt.axvline(x=cur_env.bound[1], c="green")
    plt.axhline(y=cur_env.bound[2], c="green")
    plt.axhline(y=cur_env.bound[3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")
    plt.show()
    return reward

def display_exploration_expolitation(cur_env, combined_trajectory, eps_period):
    episode_to_show = 4999
    episode_steps = combined_trajectory[episode_to_show]
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    selfLabel = 0
    # # draw link towards destination for all drones, destination for each drone didn't change
    for agentIdx in cur_env.all_agents:
        plt.plot([cur_env.all_agents[agentIdx].ini_pos[0], cur_env.all_agents[agentIdx].goal[0][0]],
                 [cur_env.all_agents[agentIdx].ini_pos[1], cur_env.all_agents[agentIdx].goal[0][1]], '--', color='c')
    x_explore = []
    y_explore = []
    x_exploit = []
    y_exploit = []
    for epsIDX, episode_steps in enumerate(combined_trajectory):
        for step_idx, agents_traj in enumerate(episode_steps):
            for ea_idx, each_agent in enumerate(agents_traj):
                if epsIDX <= eps_period:
                    x_explore.append(each_agent[0])
                    y_explore.append(each_agent[1])
                else:
                    x_exploit.append(each_agent[0])
                    y_exploit.append(each_agent[1])

                # plot self_circle of the drone
                if selfLabel == 0:
                    self_circle = Point(each_agent[0], each_agent[1]).buffer(2.5, cap_style='round')
                    grid_mat_Scir = shapelypoly_to_matpoly(self_circle, False, 'k')
                    ax.add_patch(grid_mat_Scir)

                    # check transition
                    # map output action from NN to actual range

                    coe_a = 4
                    timestep = 1
                    ax_, ay_ = 1, 1
                    ax_ = map_range(ax_, coe_a)
                    ay_ = map_range(ay_, coe_a)
                    curVelx = cur_env.all_agents[ea_idx].vel[0] + ax_ * timestep
                    curVely = cur_env.all_agents[ea_idx].vel[1] + ay_ * timestep
                    delta_x = curVelx * timestep
                    delta_y = curVely * timestep
                    cur_env.all_agents[ea_idx].pos = np.array([cur_env.all_agents[ea_idx].pos[0] + delta_x,
                                                               cur_env.all_agents[ea_idx].pos[1] + delta_y])
                    # plt.scatter(cur_env.all_agents[ea_idx].pos[0], cur_env.all_agents[ea_idx].pos[1], color='lightblue')




            selfLabel = 1



    # draw occupied_poly
    for one_poly in cur_env.world_map_2D_polyList[0][0]:
        one_poly_mat = shapelypoly_to_matpoly(one_poly, True, 'w', 'g')  # 4th parameter is the face color
        # ax.add_patch(one_poly_mat)
    # draw non-occupied_poly
    for zero_poly in cur_env.world_map_2D_polyList[0][1]:
        zero_poly_mat = shapelypoly_to_matpoly(zero_poly, False, 'w')
        # ax.add_patch(zero_poly_mat)

    # show building obstacles
    for poly in cur_env.buildingPolygons:
        matp_poly = shapelypoly_to_matpoly(poly, False, 'red')  # the 3rd parameter is the edge color
        ax.add_patch(matp_poly)

    cmap_exploit = colors.LinearSegmentedColormap.from_list("", ["white", "yellow"])
    cmap_explore = colors.LinearSegmentedColormap.from_list("", ["white", "blue"])

    # Hexbin exploit(Yellow) with white color for zero count
    # hb1 = plt.hexbin(np.array(x_exploit), np.array(y_exploit), gridsize=25, cmap=cmap_exploit, mincnt=1, alpha=1)
    # cb1 = plt.colorbar(hb1)
    # cb1.set_label('Exploit Hexbin')

    # # Hexbin explore(blue) with white color for zero count
    hb2 = plt.hexbin(np.array(x_explore), np.array(y_explore), gridsize=50, cmap=cmap_explore, mincnt=1, alpha=1)
    cb2 = plt.colorbar(hb2)
    cb2.set_label('Explore Hexbin')

    plt.axis('equal')
    plt.xlim(cur_env.bound[0], cur_env.bound[1])
    plt.ylim(cur_env.bound[2], cur_env.bound[3])
    plt.axvline(x=cur_env.bound[0], c="green")
    plt.axvline(x=cur_env.bound[1], c="green")
    plt.axhline(y=cur_env.bound[2], c="green")
    plt.axhline(y=cur_env.bound[3], c="green")
    plt.xlabel("X axis")
    plt.ylabel("Y axis")
    plt.show()


def action_selection_statistics(action_selection_collection):
    all_action_collection_x = []
    all_action_collection_y = []
    os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
    matplotlib.use('TkAgg')
    fig, ax = plt.subplots(1, 1)
    for each_eps_collection in action_selection_collection:
        for each_step in each_eps_collection:
            for agent_idx, agent_val in each_step.items():
                all_action_collection_x.append(agent_val[0][0])
                all_action_collection_y.append(agent_val[0][1])

    # Set the number of bins for x and y
    num_bins = 20
    # Create the 2D histogram
    plt.hist2d(all_action_collection_x, all_action_collection_y, bins=num_bins)

    # Set the x-axis and y-axis labels
    plt.xlabel('X')
    plt.ylabel('Y')

    # Set the title of the histogram
    plt.title('2D Histogram of X and Y')

    # Add a colorbar
    plt.colorbar()

    # Show the histogram
    plt.show()





